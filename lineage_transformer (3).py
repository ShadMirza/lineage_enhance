"""
================================================================================
  Enterprise Column Lineage Transformation Extractor
  ────────────────────────────────────────────────────────────────────────────
  Author  : Expert Data Engineering Team
  Version : 7.0.0 (Production)
  Python  : 3.9+

  Purpose
  -------
  Enhances a column-level lineage CSV (~180k rows) by traversing Teradata .sql
  files and extracting the *exact* mathematical / functional transformation
  logic applied between a source column and a target column.

  Outputs
  -------
  1. Enterprise_Fact_Lineage.xlsx    – original columns + transformation_logic
                                       + query_uuid
                                       Rows where sqlglot failed to parse will
                                       have transformation_logic = "PARSE_FAILED"
                                       and a query_uuid you can look up in the
                                       Dim Queries sheet for the raw SQL.
  2. Enterprise_Dim_Queries.xlsx     – query_uuid · sql_query · file_path
                                       · parse_status · parse_error
                                       Every statement gets a UUID — including
                                       ones that failed to parse — so PARSE_FAILED
                                       rows in the Fact sheet always have a
                                       traceable query_uuid here.

  Technology Stack
  ----------------
  pandas · sqlglot (dialect='teradata') · concurrent.futures.ProcessPoolExecutor
  openpyxl · re · uuid · logging

  v7.0 Changes  (consolidated from real-world Excel inspection)
  ────────────────────────────────────────────────────────────────────────────
  7. READ ROWS WITH EMPTY TARGET COLUMN
     SELECT statements that read a source column for filtering, joining, or
     aggregation no longer get tagged MISSING_TARGET_COLUMN. The new
     _extract_read_context() function inspects the parsed AST and returns:
       "Read in SELECT projection: <expr>"
       "Read in WHERE filter: <predicate>"
       "Read in JOIN ON condition: <on>"
       "Read in GROUP BY clause"  /  "Read in ORDER BY clause"
     The matching statement's UUID is linked into Dim Queries.

  8. WILDCARD '*' HANDLING
     Source or target column = '*' produces a meaningful description like
     "Wildcard SELECT * — all source columns" with the matching statement
     UUID, rather than landing in TRANSFORM_NOT_FOUND.

  9. STATIC CONSISTENCY
     When a row has source_column='STATIC VALUE' and the primary scan finds
     no candidate, a fallback now scans every parseable statement for ANY
     literal assigned to target_column. This eliminates the inconsistency
     where some STATIC rows showed extracted values ('Y', 99, CURRENT_DATE)
     while others showed only the bare word "STATIC".

 10. UNIVERSAL UUID COVERAGE
     EVERY output row (including MISSING_*, STATIC, wildcard, and parse-
     failure rows) now carries a query_uuid pointing to a real entry in
     Enterprise_Dim_Queries.xlsx. Default = first parseable statement in
     the file, so analysts can always click through to the source SQL.

 11. ORIGINAL FILE PATH PRESERVED IN OUTPUT
     The CSV's relative parent path (e.g.
     "ActiveInventory_Modified/.../file.sql") is preserved exactly in both
     output Excel files. The resolved absolute path is used for I/O only.
  ────────────────────────────────────────────────────────────────────────────
  6. NULL LITERAL UUID RESOLUTION
     Previously rows with source_column="NULL" or target_column="NULL" got
     transformation_logic="NULL" but query_uuid="" — leaving no trace of
     which statement the NULL came from.
     Now the NULL literal gate scans the file's statements to find the one
     that assigns to the target column, and links its UUID.  Fallback chain:
       (a) statement whose AST maps target_col to any expression  → use it
       (b) first parseable statement in the file                  → use it
       (c) first statement in the file (parse_ok or not)          → use it
     The NULL still lives in Dim Queries as any other statement does.
  ────────────────────────────────────────────────────────────────────────────
  5. PARSE FAILURE TRACKING
     When sqlglot cannot parse a statement the statement still receives a UUID
     and its raw SQL is stored in Enterprise_Dim_Queries.xlsx.
     Any lineage row whose transformation could not be extracted because of a
     parse failure gets:
       transformation_logic = "PARSE_FAILED"
       query_uuid           = UUID of the statement that failed to parse
     You can cross-reference the UUID in Dim Queries to see the raw SQL and
     the exact sqlglot error that caused the failure.
     Enterprise_Dim_Queries.xlsx gains two new columns:
       parse_status  – "OK" | "PARSE_FAILED"
       parse_error   – blank when OK, exact exception message when failed
================================================================================
"""

from __future__ import annotations

import logging
import os
import re
import signal
import sys
import threading
import uuid
import warnings
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed, TimeoutError as FutureTimeoutError
from multiprocessing import freeze_support
from pathlib import Path
from typing import Any, NamedTuple

import pandas as pd
import sqlglot
import sqlglot.expressions as exp
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# Raise Python's recursion limit so deeply nested SQL (e.g. 10-level
# correlated subqueries) doesn't hit the default 1000-frame cap.
sys.setrecursionlimit(5000)

# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("lineage_extractor.log", mode="w", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Constants / Sentinels
# ──────────────────────────────────────────────────────────────────────────────
MISSING_SOURCE   = "MISSING_SOURCE_COLUMN"
MISSING_TARGET   = "MISSING_TARGET_COLUMN"
MISSING_BOTH     = "MISSING_BOTH_COLUMNS"
TRANSFORM_NF     = "TRANSFORM_NOT_FOUND"
PARSE_FAILED     = "PARSE_FAILED"          # sqlglot could not build an AST
DIRECT_PASSTHRU  = "Direct pass through"
STATIC_MARKERS = {
    # ── all variants observed in Teradata lineage CSVs ────────────────────────
    "*",                   # wildcard / positional
    "STATIC",              # bare word
    "STATICVALUE",         # no space       (seen in your CSV, rows 3 & 14)
    "STATIC VALUE",        # with space     (seen in your CSV, rows 22-24)
    "STATIC_VALUE",        # underscore
    "HARDCODED",           # explicit label
    "HARDCODED_VALUE",
    "HARDCODED VALUE",
    "LITERAL",
    "LITERAL_VALUE",
    "LITERAL VALUE",
    "CONSTANT",
    "CONSTANT_VALUE",
    "CONSTANT VALUE",
    "DERIVED",             # some lineage tools emit this for computed statics
}
STATIC_FALLBACK = "STATIC"

# ── Hardcoded SQL NULL literal ────────────────────────────────────────────────
# When the CSV has the string "NULL" in source_column or target_column it means
# the SQL is writing a literal NULL value — NOT that the column is unknown.
# Handled before the null-safety gate so it never becomes MISSING_*.
NULL_LITERAL = "NULL"

# ──────────────────────────────────────────────────────────────────────────────
# Execution-safety limits  (prevent hangs / runaway processing)
# ──────────────────────────────────────────────────────────────────────────────
# Maximum wall-clock seconds a single worker subprocess may run before the
# main process cancels it and marks all its rows TRANSFORM_NOT_FOUND.
WORKER_TIMEOUT_SEC: int     = 300        # 5 minutes per SQL file

# Maximum wall-clock seconds allowed for one sqlglot.parse() call.
# Pathological SQL (e.g. a 50,000-token single expression) can cause
# sqlglot to spend minutes building an AST.
STMT_PARSE_TIMEOUT_SEC: int = 30         # 30 seconds per statement

# Statements longer than this (chars) are skipped entirely — sqlglot would
# likely time out anyway and they are almost certainly data dumps, not DML.
MAX_STMT_LENGTH_CHARS: int  = 200_000    # 200 KB per statement

# If a file splits into more than this many statements, process only the
# first N and warn.  Guards against files with millions of inline VALUES rows.
MAX_STMTS_PER_FILE: int     = 500
_VAR_PATTERN  = re.compile(r"\$\{?([A-Za-z_][A-Za-z0-9_]*)\}?\.([A-Za-z_][A-Za-z0-9_]*)")
_VAR_PLACEHOLDER_TPL = "__SHELLVAR_{idx}__"
_PLACEHOLDER_RE      = re.compile(r"__SHELLVAR_(\d+)__")

# Strip ALL qualifier levels from a column reference:
#   col                  →  col
#   t.col                →  col      (single alias)
#   schema.t.col         →  col      (two levels)
#   db.schema.t.col      →  col      (three levels)
_ALL_QUALIFIERS_STRIP = re.compile(r"^(?:[A-Za-z_][A-Za-z0-9_]*\.)+", re.IGNORECASE)


# ──────────────────────────────────────────────────────────────────────────────
# Pre-processor: replace Teradata shell variables with safe mock identifiers
# ──────────────────────────────────────────────────────────────────────────────

def _replace_shell_vars(sql: str) -> tuple[str, dict[int, str]]:
    """
    Replaces `$DB.table` / `${DB}.table` with deterministic mock identifiers
    so that sqlglot can parse the statement without choking on `$`.

    Returns
    -------
    cleaned_sql : str
        SQL with all shell-variable references substituted.
    mapping     : dict[int, str]
        Index → original token mapping for reversal if needed.
    """
    mapping: dict[int, str] = {}
    counter = 0

    def _replacer(m: re.Match) -> str:
        nonlocal counter
        original = m.group(0)
        mapping[counter] = original
        placeholder = _VAR_PLACEHOLDER_TPL.format(idx=counter)
        counter += 1
        return placeholder

    cleaned = _VAR_PATTERN.sub(_replacer, sql)
    return cleaned, mapping


def _restore_shell_vars(text: str, mapping: dict[int, str]) -> str:
    """Reverse substitution: restore original `$DB.table` tokens."""
    def _restorer(m: re.Match) -> str:
        idx = int(m.group(1))
        return mapping.get(idx, m.group(0))
    return _PLACEHOLDER_RE.sub(_restorer, text)


# ──────────────────────────────────────────────────────────────────────────────
# SQL Statement Splitter
# ──────────────────────────────────────────────────────────────────────────────

def split_sql_statements(raw_sql: str) -> list[str]:
    """
    Naively splits a SQL file into individual statements on semicolons,
    strips comment-only blocks, and returns non-empty strings.

    Note: Teradata uses `;` as the standard statement terminator.

    The block-comment regex uses a POSIX-safe character-class pattern instead
    of `.*?` with DOTALL.  The `.*?` + DOTALL combination can trigger
    catastrophic backtracking on malformed SQL that has an unclosed `/*`
    (the engine backtracks exponentially looking for a `*/` that never comes).
    """
    no_sl_comments = re.sub(r"--[^\n]*", "", raw_sql)
    # Safe block-comment remover — no catastrophic backtracking on unclosed /*
    no_comments    = re.sub(r"/\*[^*]*\*+(?:[^*/][^*]*\*+)*/", "", no_sl_comments)

    return [s.strip() for s in no_comments.split(";") if s.strip()]


def _safe_parse(stmt_text: str) -> list[exp.Expression]:
    """
    Call sqlglot.parse() with a hard wall-clock timeout so a single
    pathological statement can never freeze the worker subprocess.

    Strategy
    --------
    * On POSIX (Linux / macOS): use ``signal.alarm`` — zero overhead,
      guaranteed delivery even if the C extension holding the GIL.
    * On Windows (no SIGALRM): run the parse in a daemon thread and join
      with a timeout.  If it exceeds the limit the thread is abandoned
      (Python cannot kill a thread, but the daemon flag ensures it won't
      prevent process exit) and a TimeoutError is raised.

    Raises
    ------
    TimeoutError   – parse did not finish within STMT_PARSE_TIMEOUT_SEC.
    Exception      – any sqlglot parse error, re-raised as-is.
    """
    timeout = STMT_PARSE_TIMEOUT_SEC

    # ── POSIX path ────────────────────────────────────────────────────────────
    if hasattr(signal, "SIGALRM"):
        def _alarm_handler(signum, frame):
            raise TimeoutError(
                f"sqlglot.parse timed out after {timeout}s "
                f"(statement length: {len(stmt_text)} chars)"
            )
        old_handler = signal.signal(signal.SIGALRM, _alarm_handler)
        signal.alarm(timeout)
        try:
            return sqlglot.parse(
                stmt_text, read="teradata",
                error_level=sqlglot.ErrorLevel.RAISE,
            )
        finally:
            signal.alarm(0)
            signal.signal(signal.SIGALRM, old_handler)

    # ── Windows / no SIGALRM path ─────────────────────────────────────────────
    result_holder: list[Any]    = []
    error_holder:  list[BaseException] = []

    def _parse_thread() -> None:
        try:
            result_holder.append(
                sqlglot.parse(
                    stmt_text, read="teradata",
                    error_level=sqlglot.ErrorLevel.RAISE,
                )
            )
        except BaseException as exc:        # noqa: BLE001
            error_holder.append(exc)

    t = threading.Thread(target=_parse_thread, daemon=True)
    t.start()
    t.join(timeout=timeout)

    if t.is_alive():
        # Thread is still running — we cannot kill it, but it's a daemon so
        # it won't block process exit.  Raise so the caller treats it as a
        # parse failure.
        raise TimeoutError(
            f"sqlglot.parse timed out after {timeout}s "
            f"(statement length: {len(stmt_text)} chars)"
        )
    if error_holder:
        raise error_holder[0]
    return result_holder[0] if result_holder else []


# ──────────────────────────────────────────────────────────────────────────────
# AST Extraction Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _unqualified(name: str) -> str:
    """
    Strip ALL table/schema/db qualifier prefixes from a column reference.

    Examples
    --------
    _unqualified("customer_id")             → "customer_id"
    _unqualified("t.customer_id")           → "customer_id"
    _unqualified("schema.t.customer_id")    → "customer_id"
    _unqualified("db.schema.t.customer_id") → "customer_id"
    """
    return _ALL_QUALIFIERS_STRIP.sub("", name.strip()).strip()


def _is_passthrough(expr_sql: str, src_col: str) -> bool:
    """
    Return True when expr_sql is purely a (possibly multi-qualified) reference
    to src_col — i.e. no operators, functions, or casts involved.

    The CSV may omit the alias prefix (source_column = CUSTOMER_ID) while the
    SQL writes t.customer_id or db.schema.t.customer_id.  After stripping all
    qualifier levels both sides must match exactly (case-insensitive).

    Also handles the case where sqlglot emits the column with its original
    mixed-case (e.g. "customer_Id") by uppercasing both sides.
    """
    bare = _unqualified(expr_sql).upper()
    return bare == src_col.upper()


def _expr_to_sql(node: exp.Expression) -> str:
    """Render an AST expression node back to its SQL string."""
    try:
        return node.sql(dialect="teradata").strip()
    except Exception:
        return str(node).strip()


def _normalize_col(col: Any) -> str:
    """
    Coerce a column value to a clean upper-case string.

    Returns
    -------
    ""      – blank cells, pandas NaN, "nan", "none", "na", "n/a"
    "NULL"  – preserved: means a hardcoded SQL NULL literal in the query
    UPPER   – all other non-empty values (e.g. "customer_id" → "CUSTOMER_ID")

    Note: "NULL" is intentionally NOT collapsed to "" here. The worker's
    null-safety gate checks for NULL_LITERAL before checking for empty string,
    so a source_column or target_column of "NULL" gets the correct treatment
    (transformation_logic = "NULL") rather than MISSING_SOURCE/TARGET.
    """
    if col is None:
        return ""
    s = str(col).strip()
    # Only standard NA markers become empty — "NULL" is preserved as "NULL"
    if s.lower() in ("nan", "none", "", "na", "n/a", "#n/a"):
        return ""
    return s.upper()


def _bare_table(table_ref: str) -> str:
    """
    Strip database/schema prefix from a table reference so we can compare
    CSV table names against SQL statement targets.

    Handles all common forms seen in real-world Teradata lineage CSVs:

    Examples
    --------
    _bare_table("$CSMIDB.P4451_CMR_MARKER")     → "P4451_CMR_MARKER"
    _bare_table("${CSMIDB}.P4451_CMR_MARKER")   → "P4451_CMR_MARKER"
    _bare_table("$DWDB.fact_orders")            → "FACT_ORDERS"
    _bare_table("fact_orders")                  → "FACT_ORDERS"
    _bare_table("db.schema.fact_orders")        → "FACT_ORDERS"
    _bare_table('"DWDB"."fact_orders"')         → "FACT_ORDERS"
    _bare_table("  P4451_CMR_MARKER  ")         → "P4451_CMR_MARKER"
    """
    if not table_ref:
        return ""
    s = str(table_ref).strip()
    if not s:
        return ""
    # Remove surrounding/embedded double-quotes that some dialects use
    s = s.replace('"', "")
    # Remove leading $VAR. or ${VAR}. prefix
    s = re.sub(r"^\$\{?[A-Za-z_][A-Za-z0-9_]*\}?\.", "", s)
    # Strip any remaining db.schema. qualifiers (one or more levels)
    s = _ALL_QUALIFIERS_STRIP.sub("", s)
    return s.upper()


def _table_matches(csv_table: str, sql_table: str) -> bool:
    """
    Forgiving comparison between a table name from the CSV and one from
    the parsed SQL. Both are normalised through _bare_table so that:
        $CSMIDB.P4451_CMR_MARKER  matches  P4451_CMR_MARKER
        DWDB.fact_orders          matches  fact_orders
        "fact_orders"             matches  fact_orders

    Returns False if either side is empty (no info to match on).
    """
    a, b = _bare_table(csv_table), _bare_table(sql_table)
    return bool(a) and bool(b) and a == b


def _extract_stmt_target_table(parsed: exp.Expression) -> str:
    """
    Return the bare (unqualified, upper-case) table name that this DML
    statement writes INTO, so we can match it against target_table from CSV.

    CRITICAL: We MUST use `node.this` to get the write target, NOT
    `parsed.find(exp.Table)`. The latter walks the entire tree and grabs
    whichever Table node it finds first — which for an INSERT...SELECT
    or UPDATE...FROM is usually a SOURCE table from the SELECT/JOIN, not
    the actual write target. That's the root cause of every row coming
    back as TRANSFORM_NOT_FOUND — the priority filter was matching against
    the wrong table on every statement.

    For INSERT INTO tbl (cols)... the .this is exp.Schema, whose .this
    is the Table node.
    For UPDATE tbl SET... the .this IS the Table directly.
    For MERGE INTO tbl... the .this IS the Table directly.

    Returns "" if the statement has no identifiable write target.
    """
    try:
        # INSERT
        insert = parsed.find(exp.Insert)
        if insert is not None:
            target = insert.this
            # INSERT INTO tbl (col1, col2)  →  Schema node wrapping Table
            if isinstance(target, exp.Schema):
                target = target.this
            if isinstance(target, exp.Table):
                return _bare_table(target.name or "")

        # UPDATE
        update = parsed.find(exp.Update)
        if update is not None:
            target = update.this
            if isinstance(target, exp.Table):
                return _bare_table(target.name or "")

        # MERGE
        merge = parsed.find(exp.Merge)
        if merge is not None:
            target = merge.this
            if isinstance(target, exp.Table):
                return _bare_table(target.name or "")
    except Exception:
        pass
    return ""


def _resolve_source_alias(parsed: exp.Expression, source_table: str) -> str | None:
    """
    Walk the FROM / JOIN clauses of the parsed statement and return the alias
    used for `source_table`.

    CRITICAL: In sqlglot, TABLE aliases are NOT wrapped in exp.Alias nodes
    (those are for COLUMN aliases like `expr AS col_name`). A Table alias
    is stored directly on the Table node — accessed via `.alias` (string)
    or `.alias_or_name`. Iterating exp.Alias nodes for table aliases is
    a bug — that's why src_alias was always None and pass-through detection
    silently degraded.

    Parameters
    ----------
    source_table : bare (unqualified) table name from the CSV, upper-case.

    Returns
    -------
    The alias string (upper-case) if found and aliased, or the bare table
    name itself if the table appears without an alias, or None if not found.
    """
    if not source_table:
        return None
    try:
        for tbl_node in parsed.find_all(exp.Table):
            if _bare_table(tbl_node.name or "") == source_table:
                # Table.alias returns the alias string, or "" if no alias.
                # Cast through str() to be safe against any unexpected types.
                alias_str = str(tbl_node.alias) if tbl_node.alias else ""
                if alias_str:
                    return alias_str.upper()
                # No alias — caller can still use the bare table name itself
                return source_table
    except Exception:
        pass
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Core Transformation Extractor  (per statement, per column pair)
# ──────────────────────────────────────────────────────────────────────────────

def _extract_from_select(
    parsed: exp.Expression,
    src_col: str,
    tgt_col: str,
) -> str | None:
    """
    Handle SELECT projections (plain SELECT or CREATE TABLE AS SELECT).
    Finds the SELECT expression whose alias (or implicit name) matches
    `tgt_col`, then returns the rendered expression.

    For INSERT … SELECT, this is called after the caller resolves positional
    mapping (see `_extract_from_insert`).
    """
    selects = parsed.find_all(exp.Select)
    for sel in selects:
        for projection in sel.expressions:
            # Resolve alias
            alias: str | None = None
            if isinstance(projection, exp.Alias):
                alias = projection.alias.upper()
                inner = projection.this
            else:
                inner = projection
                # Try to get the column name from a bare Column node
                if isinstance(inner, exp.Column):
                    alias = inner.name.upper() if inner.name else None
                elif isinstance(inner, exp.Star):
                    alias = "*"
                else:
                    alias = None

            if alias and alias == tgt_col:
                logic = _expr_to_sql(inner)
                # Flexible pass-through: strip ALL qualifier levels before comparing.
                # CSV has alias-free names (e.g. CUSTOMER_ID) but SQL may write
                # t.customer_id, schema.t.customer_id, etc.
                if _is_passthrough(logic, src_col):
                    return DIRECT_PASSTHRU
                return logic
    return None


def _extract_from_insert(
    parsed: exp.Expression,
    src_col: str,
    tgt_col: str,
) -> str | None:
    """
    Handle INSERT INTO target (col1, col2, …) SELECT expr1, expr2, …
    Maps target columns by positional index to SELECT expressions.
    """
    insert = parsed.find(exp.Insert)
    if insert is None:
        return None

    # Collect explicit column list from INSERT INTO tbl (c1, c2, …)
    tgt_columns: list[str] = []
    schema = insert.find(exp.Schema)
    if schema:
        for col in schema.find_all(exp.Column):
            tgt_columns.append(col.name.upper())

    # Find the SELECT part
    sel = insert.find(exp.Select)
    if sel is None:
        return None

    select_exprs = sel.expressions

    if tgt_columns:
        # Named INSERT – match by name
        if tgt_col not in tgt_columns:
            return None
        idx = tgt_columns.index(tgt_col)
        if idx >= len(select_exprs):
            return None
        logic = _expr_to_sql(select_exprs[idx])
        if _is_passthrough(logic, src_col):
            return DIRECT_PASSTHRU
        return logic
    else:
        # Positional INSERT (no explicit column list) – fallback to SELECT walk
        return _extract_from_select(parsed, src_col, tgt_col)


def _extract_from_update(
    parsed: exp.Expression,
    src_col: str,
    tgt_col: str,
) -> str | None:
    """
    Handle UPDATE tbl SET col = expr …
    Finds the EQ node whose left side matches `tgt_col`.
    """
    update = parsed.find(exp.Update)
    if update is None:
        return None

    for eq in parsed.find_all(exp.EQ):
        left = eq.left
        if isinstance(left, exp.Column) and left.name.upper() == tgt_col:
            logic = _expr_to_sql(eq.right)
            if _is_passthrough(logic, src_col):
                return DIRECT_PASSTHRU
            return logic
    return None


def _extract_from_merge(
    parsed: exp.Expression,
    src_col: str,
    tgt_col: str,
) -> str | None:
    """
    Handle MERGE INTO tbl USING … ON …
      WHEN MATCHED THEN UPDATE SET col = expr
      WHEN NOT MATCHED THEN INSERT (cols) VALUES (exprs)

    Both branches are inspected.  If both yield a result they are joined
    with ' | '.
    """
    merge = parsed.find(exp.Merge)
    if merge is None:
        return None

    results: list[str] = []

    for when in parsed.find_all(exp.When):
        # ── MATCHED → UPDATE SET ─────────────────────────────────────────────
        for eq in when.find_all(exp.EQ):
            left = eq.left
            if isinstance(left, exp.Column) and left.name.upper() == tgt_col:
                logic = _expr_to_sql(eq.right)
                results.append(DIRECT_PASSTHRU if _is_passthrough(logic, src_col) else logic)

        # ── NOT MATCHED → INSERT (cols) VALUES (exprs) ───────────────────────
        schema = when.find(exp.Schema)
        values_node = when.find(exp.Tuple)

        if schema and values_node:
            ins_cols = [c.name.upper() for c in schema.find_all(exp.Column)]
            ins_vals = list(values_node.expressions)
            if tgt_col in ins_cols:
                idx = ins_cols.index(tgt_col)
                if idx < len(ins_vals):
                    logic = _expr_to_sql(ins_vals[idx])
                    results.append(DIRECT_PASSTHRU if _is_passthrough(logic, src_col) else logic)

    return " | ".join(dict.fromkeys(results)) if results else None   # dedup, order-preserving


def _is_literal_expr(expr_sql: str) -> bool:
    """
    Return True when expr_sql is a SQL literal — quoted string, number,
    CURRENT_DATE/TIMESTAMP, or NULL — i.e. a hardcoded value rather than
    a column reference or function that reads from a source table.

    Used by the static bypass to prefer literal assignments (like 'NO', 99,
    CURRENT_DATE) over column-reference assignments (like GT4451_CMR.CMRMARKER)
    when multiple statements in the same file assign to the same target column.
    """
    s = expr_sql.strip().upper()
    if not s:
        return False
    # Quoted string literals: 'value'
    if s.startswith("'") and s.endswith("'"):
        return True
    # Numeric literals (integer or decimal)
    if re.fullmatch(r"-?\d+(\.\d+)?", s):
        return True
    # Known SQL date/time constants
    if s in ("CURRENT_DATE", "CURRENT_TIME", "CURRENT_TIMESTAMP",
             "DATE", "TIME", "TIMESTAMP", "NULL"):
        return True
    # CAST of a literal: CAST('value' AS TYPE) — still a literal
    if s.startswith("CAST(") and re.search(r"CAST\s*\(\s*'", s, re.I):
        return True
    return False


class _CandidateMatch(NamedTuple):
    """
    A single candidate result from one parsed statement, with full
    source-and-target table context baked in.

    score meaning
    ─────────────
    3  –  statement targets tgt_tbl  AND  expression references src_tbl
           → highest confidence; both sides of the lineage confirmed
    2  –  statement targets tgt_tbl  but  expression source unverified
           → correct write target, source ambiguous (e.g. bare column, no alias)
    1  –  statement does NOT target tgt_tbl (fallback only)
           → lowest confidence; used only when nothing better is found
    """
    logic:      str
    uuid:       str
    score:      int    # 1 | 2 | 3
    is_literal: bool   # True for 'NO', 99, CURRENT_DATE etc.


def _expr_refs_source(
    expr:       str,
    src_tbl:    str,
    src_alias:  str | None,
    is_static:  bool = False,
) -> bool:
    """
    Return True when `expr` appears to read from `src_tbl`.

    Rules:
    1. Literals ('NO', 99, CURRENT_DATE) — accept only for STATIC rows.
       For non-static rows a literal means the SQL is hardcoding a value,
       not reading from the source column — so return False (score stays 2).
    2. alias.col in expr    →  True (confirmed)
    3. src_tbl.col in expr  →  True (confirmed, table used without alias)
    4. No qualifiers at all  →  True (bare col, ambiguous but accepted)
    5. All qualifiers are different from src_tbl/alias → False (wrong table)
    """
    if not src_tbl:
        return True

    if _is_literal_expr(expr):
        return is_static   # literals belong to static rows only

    su = expr.upper()

    if src_alias and (src_alias.upper() + ".") in su:
        return True
    if (src_tbl.upper() + ".") in su:
        return True

    qualifiers = {m.group(1).upper() for m in re.finditer(
        r"\b([A-Za-z_][A-Za-z0-9_]*)\.", su
    )}

    if not qualifiers:
        return True    # bare column — ambiguous, accept (score will be 2)

    accepted = {src_tbl.upper()}
    if src_alias:
        accepted.add(src_alias.upper())
    if qualifiers.isdisjoint(accepted):
        return False

    return True


def _extract_static_target_logic(
    parsed: exp.Expression,
    tgt_col: str,
) -> str:
    """
    When source_column is a STATIC marker, find the expression assigned to
    `tgt_col` in this statement without filtering by source column.

    Returns the extracted expression string, or STATIC_FALLBACK if nothing found.

    NOTE: This function examines a single parsed statement.  The caller
    (the worker scan loop) is responsible for scanning multiple statements
    and selecting the best result using _is_literal_expr ranking.
    """
    if not tgt_col:
        return STATIC_FALLBACK

    # UPDATE SET
    for eq in parsed.find_all(exp.EQ):
        left = eq.left
        if isinstance(left, exp.Column) and left.name.upper() == tgt_col:
            return _expr_to_sql(eq.right)

    # INSERT positional / named
    insert = parsed.find(exp.Insert)
    if insert:
        schema = insert.find(exp.Schema)

        # Case A: INSERT INTO tbl (cols) SELECT exprs FROM ...
        sel = insert.find(exp.Select)
        if schema and sel:
            ins_cols = [c.name.upper() for c in schema.find_all(exp.Column)]
            if tgt_col in ins_cols:
                idx = ins_cols.index(tgt_col)
                exprs = sel.expressions
                if idx < len(exprs):
                    return _expr_to_sql(exprs[idx])

        # Case B: INSERT INTO tbl (cols) VALUES (literals) — no SELECT
        if schema:
            ins_cols = [c.name.upper() for c in schema.find_all(exp.Column)]
            if tgt_col in ins_cols:
                idx = ins_cols.index(tgt_col)
                # VALUES is parsed as exp.Tuple or exp.Values
                values_node = insert.find(exp.Values) or insert.find(exp.Tuple)
                if values_node:
                    # exp.Values has .expressions = list of Tuples (one per row)
                    # exp.Tuple has .expressions = list of literal expressions
                    if isinstance(values_node, exp.Values):
                        rows = values_node.expressions
                        if rows and isinstance(rows[0], exp.Tuple):
                            row_exprs = rows[0].expressions
                            if idx < len(row_exprs):
                                return _expr_to_sql(row_exprs[idx])
                    else:
                        row_exprs = values_node.expressions
                        if idx < len(row_exprs):
                            return _expr_to_sql(row_exprs[idx])

    # MERGE – check all WHEN branches
    for eq in parsed.find_all(exp.EQ):
        left = eq.left
        if isinstance(left, exp.Column) and left.name.upper() == tgt_col:
            return _expr_to_sql(eq.right)

    # SELECT alias
    for sel in parsed.find_all(exp.Select):
        for proj in sel.expressions:
            if isinstance(proj, exp.Alias):
                if proj.alias.upper() == tgt_col:
                    return _expr_to_sql(proj.this)

    return STATIC_FALLBACK


# ──────────────────────────────────────────────────────────────────────────────
# Single-statement dispatcher
# ──────────────────────────────────────────────────────────────────────────────

def _extract_read_context(parsed: exp.Expression, src_col: str) -> tuple[str, bool]:
    """
    For READ-relation rows where target_column is empty: find HOW the source
    column is used in the statement and return a human-readable context
    description.

    Returns (description, found_flag).

    Recognised contexts (in priority order):
      • SELECT projection         — col appears in SELECT list
      • WHERE filter              — col appears in WHERE predicate
      • JOIN condition            — col appears in ON / USING
      • GROUP BY / ORDER BY       — col is grouping or sorting key
      • Used as input             — generic fallback
    """
    if not src_col or parsed is None:
        return ("", False)

    SU = src_col.upper()

    # Helper: does any Column node under `node` match src_col?
    def _has_col(node) -> bool:
        try:
            for c in node.find_all(exp.Column):
                if (c.name or "").upper() == SU:
                    return True
        except Exception:
            pass
        return False

    # ── SELECT projection ─────────────────────────────────────────────────────
    try:
        for sel in parsed.find_all(exp.Select):
            for proj in sel.expressions:
                if _has_col(proj):
                    inner = proj.this if isinstance(proj, exp.Alias) else proj
                    expr_sql = _expr_to_sql(inner) or src_col
                    if isinstance(proj, exp.Alias):
                        return (f"Read in SELECT (aliased as {proj.alias.upper()}): {expr_sql}", True)
                    return (f"Read in SELECT projection: {expr_sql}", True)
    except Exception:
        pass

    # ── JOIN condition ────────────────────────────────────────────────────────
    try:
        for join in parsed.find_all(exp.Join):
            on_node = join.args.get("on")
            if on_node and _has_col(on_node):
                on_sql = _expr_to_sql(on_node)
                return (f"Read in JOIN ON condition: {on_sql}", True)
    except Exception:
        pass

    # ── WHERE clause ──────────────────────────────────────────────────────────
    try:
        for where in parsed.find_all(exp.Where):
            if _has_col(where):
                where_sql = _expr_to_sql(where.this) if where.this else _expr_to_sql(where)
                return (f"Read in WHERE filter: {where_sql}", True)
    except Exception:
        pass

    # ── GROUP BY / ORDER BY ───────────────────────────────────────────────────
    try:
        for group in parsed.find_all(exp.Group):
            if _has_col(group):
                return (f"Read in GROUP BY clause", True)
        for order in parsed.find_all(exp.Order):
            if _has_col(order):
                return (f"Read in ORDER BY clause", True)
    except Exception:
        pass

    # ── Generic fallback: column appears somewhere in the statement ──────────
    if _has_col(parsed):
        return (f"Read as input column", True)

    return ("", False)


def _extract_wildcard_context(parsed: exp.Expression) -> tuple[str, bool]:
    """
    Handle source_column='*' or target_column='*'. The wildcard means
    "all columns" — produce a human-readable description of what the
    statement does.

    Returns (description, found_flag).
    """
    if parsed is None:
        return ("", False)

    try:
        # SELECT *  or  SELECT t.*
        for sel in parsed.find_all(exp.Select):
            for proj in sel.expressions:
                if isinstance(proj, exp.Star):
                    return ("Wildcard SELECT * — all source columns", True)
                if isinstance(proj, exp.Column) and isinstance(proj.this, exp.Star):
                    return (f"Wildcard SELECT {proj.table}.* — all columns from {proj.table}", True)

        # INSERT INTO tbl SELECT *  (positional copy of all source columns)
        ins = parsed.find(exp.Insert)
        if ins:
            sel = ins.find(exp.Select)
            if sel:
                for proj in sel.expressions:
                    if isinstance(proj, exp.Star):
                        return ("Wildcard INSERT ... SELECT * — copy all columns", True)

        # DELETE / UPDATE without column list — affects all columns
        if parsed.find(exp.Delete):
            return ("Wildcard DELETE — affects all rows", True)
    except Exception:
        pass

    return ("", False)


def extract_transformation(
    parsed: exp.Expression,
    src_col: str,
    tgt_col: str,
    is_static: bool,
    src_alias: str | None = None,
) -> str:
    """
    Dispatch to the correct extractor based on the AST root type.

    Parameters
    ----------
    parsed     : sqlglot AST root expression
    src_col    : normalised source column name (upper-case, no qualifier)
    tgt_col    : normalised target column name (upper-case, no qualifier)
    is_static  : True if the original source_column was a STATIC marker
    src_alias  : resolved alias for the source table in this statement,
                 used to tighten the pass-through check.
                 e.g. if source_table is ICV_FINAL_ACTIVE_CM and the SQL has
                 "FROM $CSMIDB.ICV_FINAL_ACTIVE_CM t", src_alias = "T".
                 When provided, only "T.col" or bare "col" are accepted as
                 pass-throughs — not "s.col" from a different joined table.
    """
    if is_static:
        return _extract_static_target_logic(parsed, tgt_col)

    result: str | None = None

    if parsed.find(exp.Merge):
        result = _extract_from_merge(parsed, src_col, tgt_col)

    if result is None and parsed.find(exp.Update):
        result = _extract_from_update(parsed, src_col, tgt_col)

    if result is None and parsed.find(exp.Insert):
        result = _extract_from_insert(parsed, src_col, tgt_col)

    if result is None:
        result = _extract_from_select(parsed, src_col, tgt_col)

    if result is None:
        return TRANSFORM_NF

    # Pass-through check — tighten using resolved alias when available
    if result != DIRECT_PASSTHRU:
        if src_alias:
            # Accept:  t.col_name  OR  bare col_name
            # Reject:  s.col_name  (different table's alias)
            bare = _unqualified(result).upper()
            qualifier = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\.", result.strip())
            if bare == src_col:
                if qualifier is None or qualifier.group(1).upper() == src_alias.upper():
                    return DIRECT_PASSTHRU
        else:
            if _is_passthrough(result, src_col):
                return DIRECT_PASSTHRU

    return result


# ──────────────────────────────────────────────────────────────────────────────
# Worker payload & file-level processor
# ──────────────────────────────────────────────────────────────────────────────

def process_single_file(task: dict[str, Any]) -> dict[str, Any]:
    """
    Worker function executed inside a subprocess.

    Parameters (inside `task`)
    --------------------------
    file_path   : str – absolute path to the .sql file
    pairs       : list[dict] – each dict has keys:
                    row_id, source_table, source_column,
                    target_table, target_column, relation

    Returns
    -------
    dict with keys:
        results  : list[dict]  – enriched row dicts (row_id + new cols)
        queries  : list[dict]  – {query_uuid, sql_query, file_path}
        errors   : list[str]   – any non-fatal warnings
    """
    file_path:      str = task["file_path"]
    file_path_orig: str = task.get("file_path_orig", file_path)
    pairs: list[dict]   = task["pairs"]

    results: list[dict] = []
    queries: list[dict] = []
    errors:  list[str]  = []

    # ── 1. Read the file ──────────────────────────────────────────────────────
    try:
        raw_sql = Path(file_path).read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        errors.append(f"[READ ERROR] {file_path}: {exc}")
        for p in pairs:
            results.append({**p, "transformation_logic": TRANSFORM_NF, "query_uuid": ""})
        return {"results": results, "queries": queries, "errors": errors}

    # ── 2. Pre-process: replace shell variables ───────────────────────────────
    clean_sql, var_map = _replace_shell_vars(raw_sql)

    # ── 3. Split into statements and parse each ───────────────────────────────
    stmts = split_sql_statements(clean_sql)

    # Guard: too many statements → almost certainly a data-dump file.
    if len(stmts) > MAX_STMTS_PER_FILE:
        errors.append(
            f"[SKIP WARN] {Path(file_path).name}: {len(stmts)} statements found, "
            f"processing only the first {MAX_STMTS_PER_FILE} "
            f"(limit MAX_STMTS_PER_FILE={MAX_STMTS_PER_FILE})"
        )
        stmts = stmts[:MAX_STMTS_PER_FILE]

    stmt_catalog: list[dict] = []

    for stmt_text in stmts:
        stmt_uuid       = str(uuid.uuid4())
        restored        = _restore_shell_vars(stmt_text, var_map)
        parse_ok        = True
        parse_error_msg = ""
        parsed          = None

        # Guard: statement too large → skip to avoid hanging sqlglot.
        if len(stmt_text) > MAX_STMT_LENGTH_CHARS:
            parse_ok        = False
            parse_error_msg = (
                f"Statement skipped: {len(stmt_text):,} chars exceeds "
                f"MAX_STMT_LENGTH_CHARS={MAX_STMT_LENGTH_CHARS:,}"
            )
            errors.append(f"[SIZE SKIP] {Path(file_path).name} stmt {stmt_uuid[:8]}: {parse_error_msg}")
        else:
            try:
                parsed_list = _safe_parse(stmt_text)
                parsed = parsed_list[0] if parsed_list else None
                if parsed is None:
                    parse_ok        = False
                    parse_error_msg = "sqlglot returned empty AST (unsupported or DDL statement)"
            except TimeoutError as te:
                parse_ok        = False
                parse_error_msg = str(te)
                errors.append(f"[TIMEOUT] {Path(file_path).name} stmt {stmt_uuid[:8]}: {parse_error_msg}")
            except Exception as parse_exc:
                parse_ok        = False
                parse_error_msg = str(parse_exc)
                errors.append(f"[PARSE WARN] {Path(file_path).name} stmt {stmt_uuid[:8]}: {parse_error_msg}")

        stmt_catalog.append({
            "uuid":        stmt_uuid,
            "parsed":      parsed,
            "restored":    restored,
            "parse_ok":    parse_ok,
            "parse_error": parse_error_msg,
        })

        # Every statement — parsed or not — goes into the queries catalog
        # so it appears in Enterprise_Dim_Queries.xlsx and can be reviewed.
        # We use file_path_orig here (the original CSV string) so the output
        # shows relative paths exactly as the user provided them, not the
        # resolved absolute paths used during execution.
        queries.append({
            "query_uuid":   stmt_uuid,
            "sql_query":    restored,
            "file_path":    file_path_orig,
            "parse_status": "OK" if parse_ok else "PARSE_FAILED",
            "parse_error":  parse_error_msg,
        })

    # Pre-compute: does this file have ANY statements that failed to parse?
    failed_stmts = [s for s in stmt_catalog if not s["parse_ok"]]

    # ── 4. For each lineage pair, scan the statements ─────────────────────────
    for pair in pairs:
        raw_src     = pair.get("source_column", "")
        raw_tgt     = pair.get("target_column", "")
        src_tbl_raw = pair.get("source_table",  "")
        tgt_tbl_raw = pair.get("target_table",  "")
        relation    = str(pair.get("relation", "")).strip().upper()

        src_col = _normalize_col(raw_src)
        tgt_col = _normalize_col(raw_tgt)
        src_tbl = _bare_table(str(src_tbl_raw))
        tgt_tbl = _bare_table(str(tgt_tbl_raw))

        # Default UUID for any row that doesn't find a specific match — points
        # to the first parseable statement so EVERY row has a Dim Queries link.
        parseable_pre = [s for s in stmt_catalog if s["parse_ok"]]
        default_uuid_pre = (parseable_pre[0]["uuid"] if parseable_pre else
                            stmt_catalog[0]["uuid"] if stmt_catalog else "")

        # ── NULL literal gate ─────────────────────────────────────────────────
        if src_col == NULL_LITERAL or tgt_col == NULL_LITERAL:
            null_uuid = ""
            real_tgt  = tgt_col if tgt_col != NULL_LITERAL else ""
            if real_tgt:
                for stmt in stmt_catalog:
                    if not stmt["parse_ok"]:
                        continue
                    try:
                        expr = _extract_static_target_logic(stmt["parsed"], real_tgt)
                        if expr and expr != STATIC_FALLBACK:
                            null_uuid = stmt["uuid"]; break
                    except Exception:
                        pass
            if not null_uuid:
                null_uuid = default_uuid_pre
            results.append({**pair, "transformation_logic": NULL_LITERAL,
                             "query_uuid": null_uuid})
            continue

        # ── Wildcard '*' gate ─────────────────────────────────────────────────
        # When source_column or target_column is '*' (and not a STATIC marker
        # — STATIC_MARKERS includes '*' but that means "static value", we
        # only treat bare '*' here when relation is READ/WRITE without
        # static intent, and source AND target table are populated)
        if (raw_src == "*" or raw_tgt == "*"):
            wc_logic = ""
            wc_uuid  = default_uuid_pre
            for stmt in stmt_catalog:
                if not stmt["parse_ok"]:
                    continue
                try:
                    desc, found = _extract_wildcard_context(stmt["parsed"])
                    if found:
                        # Prefer statements that touch the right table
                        stmt_tgt = _extract_stmt_target_table(stmt["parsed"])
                        if tgt_tbl and _table_matches(tgt_tbl, stmt_tgt):
                            wc_logic, wc_uuid = desc, stmt["uuid"]
                            break
                        if not wc_logic:
                            wc_logic, wc_uuid = desc, stmt["uuid"]
                except Exception:
                    pass
            if not wc_logic:
                wc_logic = (f"Wildcard ({raw_src or raw_tgt}) — all columns from "
                            f"{src_tbl_raw or tgt_tbl_raw}")
            results.append({**pair, "transformation_logic": wc_logic,
                             "query_uuid": wc_uuid})
            continue

        # ── READ-with-empty-target gate ───────────────────────────────────────
        # READ rows describe how a source column is consumed (filter, join,
        # projection) — they don't have a target_column. Don't drop them as
        # MISSING_TARGET — extract the read context instead.
        if relation == "READ" and src_col and not tgt_col:
            rc_logic = ""
            rc_uuid  = default_uuid_pre
            for stmt in stmt_catalog:
                if not stmt["parse_ok"]:
                    continue
                try:
                    desc, found = _extract_read_context(stmt["parsed"], src_col)
                    if found:
                        rc_logic, rc_uuid = desc, stmt["uuid"]
                        break
                except Exception:
                    pass
            if not rc_logic:
                rc_logic = f"Read column '{raw_src}' (context not extracted)"
            results.append({**pair, "transformation_logic": rc_logic,
                             "query_uuid": rc_uuid})
            continue

        # ── Missing-column gates ──────────────────────────────────────────────
        # Even MISSING rows now carry the default UUID so analysts can open
        # the file in Dim Queries and inspect manually.
        if not src_col and not tgt_col:
            results.append({**pair, "transformation_logic": MISSING_BOTH,
                             "query_uuid": default_uuid_pre}); continue
        if not src_col:
            results.append({**pair, "transformation_logic": MISSING_SOURCE,
                             "query_uuid": default_uuid_pre}); continue
        if not tgt_col:
            results.append({**pair, "transformation_logic": MISSING_TARGET,
                             "query_uuid": default_uuid_pre}); continue

        is_static = src_col.upper() in {m.upper() for m in STATIC_MARKERS}
        parseable = [s for s in stmt_catalog if s["parse_ok"]]

        # Default UUID: first parseable stmt so even TRANSFORM_NOT_FOUND rows
        # always have a traceable pointer into Dim Queries.
        default_uuid = (parseable[0]["uuid"] if parseable else
                        stmt_catalog[0]["uuid"] if stmt_catalog else "")

        # ── FULL TRAVERSE: scan EVERY parseable statement ─────────────────────
        #
        # Old approach (first-match-wins) was wrong when the same column pair
        # appears in multiple statements of the same file.  We now collect ALL
        # candidates and score each one using BOTH source_table and target_table
        # context before deciding the winner.
        #
        # Scoring rubric (_CandidateMatch.score):
        #   3  – stmt targets tgt_tbl  AND  expression references src_tbl
        #   2  – stmt targets tgt_tbl  AND  expression source unverified
        #              (bare column, no qualifier — could come from src_tbl)
        #   1  – stmt does NOT target tgt_tbl (pure fallback, last resort)
        #
        # Within the same score tier, for STATIC rows literals beat col refs.
        # If multiple score-3 matches with identical logic exist they are
        # deduplicated; differing ones are joined with  " | "  (like MERGE).
        # ─────────────────────────────────────────────────────────────────────

        all_candidates: list[_CandidateMatch] = []

        for stmt in parseable:
            stmt_tgt  = _extract_stmt_target_table(stmt["parsed"])
            # Use the forgiving comparator that handles $db.tbl vs tbl etc.
            # If we can't determine the stmt target (e.g. unsupported DML),
            # still let the candidate through at score 2 — better than dropping
            # everything and giving the user TRANSFORM_NOT_FOUND for every row.
            if not tgt_tbl:
                tgt_match = True
            elif not stmt_tgt:
                tgt_match = True   # unknown stmt target → treat as ambiguous
            else:
                tgt_match = _table_matches(tgt_tbl, stmt_tgt)

            # Resolve alias for source_table in THIS statement
            src_alias: str | None = None
            if src_tbl:
                try:
                    src_alias = _resolve_source_alias(stmt["parsed"], src_tbl)
                except Exception:
                    pass

            # Extract the expression for this (src_col, tgt_col) pair
            try:
                if is_static:
                    raw_logic = _extract_static_target_logic(stmt["parsed"], tgt_col)
                else:
                    raw_logic = extract_transformation(
                        stmt["parsed"], src_col, tgt_col, False, src_alias
                    )
            except Exception as ex:
                errors.append(
                    f"[EXTRACT WARN] {Path(file_path).name} "
                    f"pair {src_col}->{tgt_col}: {ex}"
                )
                continue

            if not raw_logic or raw_logic in (TRANSFORM_NF, STATIC_FALLBACK):
                continue

            # Check whether the extracted expression actually references src_tbl
            src_confirmed = _expr_refs_source(raw_logic, src_tbl, src_alias, is_static)

            # Score this candidate
            if tgt_match and src_confirmed:
                score = 3   # both sides of lineage confirmed ✓
            elif tgt_match:
                score = 2   # right write target, source ambiguous
            else:
                score = 1   # fallback only

            all_candidates.append(_CandidateMatch(
                logic      = raw_logic,
                uuid       = stmt["uuid"],
                score      = score,
                is_literal = _is_literal_expr(raw_logic),
            ))

        # ── Select best candidate(s) ──────────────────────────────────────────
        found_logic = TRANSFORM_NF
        found_uuid  = default_uuid

        if all_candidates:
            best_score = max(c.score for c in all_candidates)
            top        = [c for c in all_candidates if c.score == best_score]

            if is_static:
                # Within the best score tier prefer literals ('NO', 99, etc.)
                # over column references (GT4451_CMR.CMRMARKER).
                lit_top = [c for c in top if c.is_literal]
                top     = lit_top or top

            # Deduplicate by logic string, preserve order
            seen: dict[str, str] = {}   # logic → uuid
            for c in top:
                if c.logic not in seen:
                    seen[c.logic] = c.uuid

            if len(seen) == 1:
                found_logic, found_uuid = next(iter(seen.items()))
            else:
                # Multiple distinct expressions at the same score tier
                # (e.g. same column set differently in two equally-scored stmts)
                # — surface all of them joined with " | " so nothing is hidden
                found_logic = " | ".join(seen.keys())
                found_uuid  = next(iter(seen.values()))   # uuid of first match

        # ── Post-scan outcome ─────────────────────────────────────────────────
        if found_logic == TRANSFORM_NF:
            if is_static:
                # STATIC consistency fix: rather than giving up with bare
                # "STATIC", scan EVERY parseable statement (priority order)
                # for ANY literal assigned to tgt_col. This guarantees that
                # if the SQL contains a hardcoded value for the target column
                # ANYWHERE in the file, we surface it — no inconsistency
                # between rows that found 'Y' and rows that just say STATIC.
                static_lit_logic = ""
                static_lit_uuid  = default_uuid
                for stmt in parseable:
                    try:
                        expr = _extract_static_target_logic(stmt["parsed"], tgt_col)
                        if expr and expr != STATIC_FALLBACK:
                            static_lit_logic = expr
                            static_lit_uuid  = stmt["uuid"]
                            if _is_literal_expr(expr):
                                break   # literal found → done
                    except Exception:
                        pass
                if static_lit_logic:
                    found_logic = static_lit_logic
                    found_uuid  = static_lit_uuid
                else:
                    found_logic = STATIC_FALLBACK
                    # Keep the default_uuid so STATIC rows are still traceable
            elif failed_stmts:
                found_logic = PARSE_FAILED
                found_uuid  = failed_stmts[0]["uuid"]

        results.append({**pair, "transformation_logic": found_logic, "query_uuid": found_uuid})

    return {"results": results, "queries": queries, "errors": errors}


# ──────────────────────────────────────────────────────────────────────────────
# Excel Formatter (openpyxl, heavily styled)
# ──────────────────────────────────────────────────────────────────────────────

# ── Palette ───────────────────────────────────────────────────────────────────
_CLR_HEADER_BG   = "1F3864"   # dark navy
_CLR_HEADER_FG   = "FFFFFF"   # white
_CLR_ALT_ROW     = "EEF2F7"   # light blue-grey
_CLR_ACCENT      = "2E75B6"   # mid-blue
_CLR_BORDER      = "BDD7EE"   # pale blue
_CLR_FOUND_BG    = "E2EFDA"   # pale green        (transform found)
_CLR_MISS_BG     = "FCE4D6"   # pale orange       (missing column)
_CLR_STATIC_BG   = "FFF2CC"   # pale yellow       (static / hardcoded)
_CLR_PASS_BG     = "DDEEFF"   # sky blue          (direct pass through)
_CLR_PARSE_BG    = "F4CCCC"   # soft red          (parse failed — manual review)


def _thin_border() -> Border:
    side = Side(style="thin", color=_CLR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _write_header_row(ws, headers: list[str], col_widths: dict[int, int]) -> None:
    header_fill = PatternFill("solid", fgColor=_CLR_HEADER_BG)
    header_font = Font(name="Calibri", bold=True, color=_CLR_HEADER_FG, size=11)
    border      = _thin_border()
    align       = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = border
        cell.alignment = align
        # Set column width (max of header length + padding, or override)
        col_widths[ci] = max(col_widths.get(ci, 0), len(h) + 4)


def _row_fill(logic: str) -> PatternFill | None:
    if not logic:
        return None
    if logic in (MISSING_BOTH, MISSING_SOURCE, MISSING_TARGET):
        return PatternFill("solid", fgColor=_CLR_MISS_BG)
    if logic == DIRECT_PASSTHRU:
        return PatternFill("solid", fgColor=_CLR_PASS_BG)
    if logic == PARSE_FAILED:
        return PatternFill("solid", fgColor=_CLR_PARSE_BG)
    if logic in (STATIC_FALLBACK,) or logic.upper().startswith("STATIC"):
        return PatternFill("solid", fgColor=_CLR_STATIC_BG)
    if logic != TRANSFORM_NF:
        return PatternFill("solid", fgColor=_CLR_FOUND_BG)
    return None


def _write_data_rows(
    ws,
    rows: list[tuple],
    headers: list[str],
    col_widths: dict[int, int],
    logic_col_idx: int | None = None,
) -> None:
    """
    Write data rows with alternating fills, border, and auto-size tracking.
    Rows that have a special `transformation_logic` value get accent fills.
    """
    alt_fill  = PatternFill("solid", fgColor=_CLR_ALT_ROW)
    norm_font = Font(name="Calibri", size=10)
    border    = _thin_border()

    for ri, row in enumerate(rows, start=2):
        is_alt    = ri % 2 == 0
        logic_val = row[logic_col_idx - 1] if logic_col_idx else None
        row_fill  = _row_fill(str(logic_val)) if logic_val else (alt_fill if is_alt else None)

        for ci, value in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font   = norm_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            if row_fill:
                cell.fill = row_fill
            elif is_alt:
                cell.fill = alt_fill

            # Track max width (cap at 80)
            val_len = min(len(str(value)) if value is not None else 0, 80)
            col_widths[ci] = max(col_widths.get(ci, 0), val_len)


def _finalise_sheet(ws, col_widths: dict[int, int], freeze_row: int = 1) -> None:
    """Apply column widths, freeze pane, auto-filter, row height."""
    ws.freeze_panes = ws.cell(row=freeze_row + 1, column=1)
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 2, 82)


def write_fact_lineage_xlsx(df: pd.DataFrame, output_path: str) -> None:
    """Write Enterprise_Fact_Lineage.xlsx with full formatting."""
    log.info("Writing Fact Lineage workbook → %s", output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Column Lineage"

    # ── Tab colour ────────────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = _CLR_ACCENT

    headers    = list(df.columns)
    col_widths: dict[int, int] = {}

    # Identify the index of transformation_logic for accent colouring
    try:
        logic_idx = headers.index("transformation_logic") + 1
    except ValueError:
        logic_idx = None

    _write_header_row(ws, headers, col_widths)

    rows = [tuple(str(v) if pd.notna(v) else "" for v in row) for row in df.itertuples(index=False)]
    _write_data_rows(ws, rows, headers, col_widths, logic_col_idx=logic_idx)
    _finalise_sheet(ws, col_widths)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    _add_summary_sheet(wb, df)

    wb.save(output_path)
    log.info("Fact Lineage workbook saved  (%d rows)", len(df))


def _add_summary_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Append a Summary tab with aggregated counts."""
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "FF0000"

    title_font  = Font(name="Calibri", bold=True, size=14, color=_CLR_HEADER_FG)
    title_fill  = PatternFill("solid", fgColor=_CLR_HEADER_BG)
    lbl_font    = Font(name="Calibri", bold=True, size=11)
    val_font    = Font(name="Calibri", size=11)
    border      = _thin_border()

    ws.merge_cells("A1:C1")
    ws["A1"].value     = "Enterprise Lineage – Transformation Summary"
    ws["A1"].font      = title_font
    ws["A1"].fill      = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    metrics = [
        ("Total Records",          len(df)),
        ("Transformations Found",  int((~df["transformation_logic"].isin(
                                        [TRANSFORM_NF, PARSE_FAILED,
                                         MISSING_BOTH, MISSING_SOURCE, MISSING_TARGET]
                                    )).sum() if "transformation_logic" in df.columns else 0)),
        ("Direct Pass Through",    int((df["transformation_logic"] == DIRECT_PASSTHRU).sum()
                                       if "transformation_logic" in df.columns else 0)),
        ("Static Value",           int(df["transformation_logic"].str.upper()
                                       .str.startswith("STATIC").sum()
                                       if "transformation_logic" in df.columns else 0)),
        ("Parse Failed (manual)",  int((df["transformation_logic"] == PARSE_FAILED).sum()
                                       if "transformation_logic" in df.columns else 0)),
        ("Missing Source Column",  int((df["transformation_logic"] == MISSING_SOURCE).sum()
                                       if "transformation_logic" in df.columns else 0)),
        ("Missing Target Column",  int((df["transformation_logic"] == MISSING_TARGET).sum()
                                       if "transformation_logic" in df.columns else 0)),
        ("Missing Both Columns",   int((df["transformation_logic"] == MISSING_BOTH).sum()
                                       if "transformation_logic" in df.columns else 0)),
        ("Transform Not Found",    int((df["transformation_logic"] == TRANSFORM_NF).sum()
                                       if "transformation_logic" in df.columns else 0)),
    ]

    for ri, (label, val) in enumerate(metrics, start=3):
        lc = ws.cell(row=ri, column=1, value=label)
        vc = ws.cell(row=ri, column=2, value=val)
        lc.font = lbl_font; lc.border = border
        vc.font = val_font;  vc.border = border
        lc.alignment = Alignment(horizontal="left")
        vc.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def write_dim_queries_xlsx(df: pd.DataFrame, output_path: str) -> None:
    """Write Enterprise_Dim_Queries.xlsx with full formatting."""
    log.info("Writing Dim Queries workbook → %s", output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "SQL Queries"
    ws.sheet_properties.tabColor = "00B050"

    headers    = list(df.columns)
    col_widths: dict[int, int] = {}

    _write_header_row(ws, headers, col_widths)

    rows = [tuple(str(v) if pd.notna(v) else "" for v in row) for row in df.itertuples(index=False)]
    _write_data_rows(ws, rows, headers, col_widths)
    _finalise_sheet(ws, col_widths)

    # ── Highlight PARSE_FAILED rows in red so they stand out ──────────────────
    try:
        status_ci = headers.index("parse_status") + 1
        fail_fill = PatternFill("solid", fgColor=_CLR_PARSE_BG)
        for ri in range(2, ws.max_row + 1):
            if ws.cell(row=ri, column=status_ci).value == "PARSE_FAILED":
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=ri, column=ci).fill = fail_fill
    except ValueError:
        pass   # parse_status column absent — skip colouring

    wb.save(output_path)
    log.info("Dim Queries workbook saved   (%d rows)", len(df))




# ──────────────────────────────────────────────────────────────────────────────
# Orchestrator
# ──────────────────────────────────────────────────────────────────────────────

def build_tasks(df: pd.DataFrame) -> list[dict[str, Any]]:
    """
    Group the dataframe by `parent` (file path) and construct the
    worker-task dictionaries.  **No DataFrame objects are passed
    to subprocesses** – only native Python dicts and lists.

    Each task dict now carries:
      file_path        : absolute resolved path used for I/O
      file_path_orig   : original CSV string — preserved for output sheets
      pairs            : list of lineage row dicts
    """
    tasks: list[dict[str, Any]] = []

    grouped = df.groupby("parent", sort=False)
    for file_path, grp in grouped:
        pairs: list[dict] = []
        # All rows in this group share the same resolved path AND the same
        # original CSV string (groupby preserved per-row identity).  Use the
        # first row's parent_original as the canonical original path.
        first_row = grp.iloc[0]
        file_path_orig = str(first_row.get("parent_original", file_path))

        for row in grp.itertuples(index=True):
            pairs.append({
                "row_id":         row.Index,
                "source_table":   getattr(row, "source_table",  ""),
                "source_column":  getattr(row, "source_column", ""),
                "target_table":   getattr(row, "target_table",  ""),
                "target_column":  getattr(row, "target_column", ""),
                "relation":       getattr(row, "relation",      ""),
            })
        tasks.append({
            "file_path":      str(file_path),
            "file_path_orig": file_path_orig,
            "pairs":          pairs,
        })

    log.info("Built %d file-level tasks from %d rows", len(tasks), len(df))
    return tasks


def run_pipeline(
    input_csv:      str,
    output_dir:     str = ".",
    sql_base_dir:   str | None = None,
    max_workers:    int | None = None,
    chunksize:      int = 4,
    worker_timeout: int = WORKER_TIMEOUT_SEC,
) -> None:
    """
    Main pipeline entry point.

    Parameters
    ----------
    input_csv      : Path to the input lineage CSV.
    output_dir     : Directory for the two output Excel files.
    sql_base_dir   : Root directory for resolving relative `parent` paths.
                     Defaults to the directory where this script lives.
    max_workers    : Number of subprocesses (default: cpu_count).
    chunksize      : ProcessPoolExecutor map chunksize.
    worker_timeout : Max seconds a single worker may run before it is
                     cancelled and its rows marked TRANSFORM_NOT_FOUND.
                     Default: WORKER_TIMEOUT_SEC (300 s).
    """
    global WORKER_TIMEOUT_SEC
    WORKER_TIMEOUT_SEC = worker_timeout
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Load CSV ──────────────────────────────────────────────────────────────
    log.info("Loading CSV: %s", input_csv)
    df = pd.read_csv(
        input_csv,
        dtype=str,
        keep_default_na=False,
        # "NULL" / "null" intentionally NOT in na_values — they are preserved
        # as the string "NULL" so _normalize_col returns "NULL" and the worker
        # treats them as hardcoded SQL NULL literals (not missing columns).
        na_values=["", "NA", "N/A", "#N/A", "None", "none"],
    )
    log.info("Loaded %d rows, columns: %s", len(df), list(df.columns))

    required_cols = {"source_table", "source_column", "target_table",
                     "target_column", "relation", "parent", "parent_type"}
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        raise ValueError(f"Input CSV is missing required columns: {missing_cols}")

    # ── Resolve relative parent paths ─────────────────────────────────────────
    # Default base = the directory where THIS script file lives.
    # That means if your CSV has  "ActiveInventory_Modified/.../file.sql"
    # and the script is at        "/project/lineage_transformer.py"
    # the resolved path becomes   "/project/ActiveInventory_Modified/.../file.sql"
    # Override with --sql-base-dir to point at a different root.
    if sql_base_dir:
        base = Path(sql_base_dir).resolve()
        log.info("Using explicit SQL base dir: %s", base)
    else:
        base = Path(__file__).resolve().parent
        log.info("Using script directory as SQL base: %s", base)

    # Preserve the ORIGINAL parent string from the CSV — that's what gets
    # written to the output Excel files (per requirement: "no need to mention
    # full path from where it was read during execution").
    # The resolved absolute path is kept in a parallel column purely for IO.
    df["parent_original"] = df["parent"].astype(str)
    df["parent"] = df["parent"].apply(
        lambda p: str((base / p).resolve()) if p and not Path(p).is_absolute() else p
    )

    # Warn on paths that still can't be found after resolution
    unique_paths = [p for p in df["parent"].dropna().unique() if p]
    missing_files = [p for p in unique_paths if not Path(p).exists()]
    if missing_files:
        log.warning(
            "%d unique parent paths NOT found on disk after resolution (first 5): %s",
            len(missing_files), missing_files[:5],
        )

    # ── Build worker tasks (no DataFrames) ────────────────────────────────────
    tasks = build_tasks(df)

    # ── Multiprocessing ───────────────────────────────────────────────────────
    all_results: list[dict] = []
    all_queries: list[dict] = []
    all_errors:  list[str]  = []

    n_workers = max_workers or max(1, os.cpu_count() or 4)
    log.info("Launching ProcessPoolExecutor with %d workers (timeout %ds per file) …",
             n_workers, WORKER_TIMEOUT_SEC)

    completed  = 0
    timed_out  = 0

    try:
        with ProcessPoolExecutor(max_workers=n_workers) as executor:
            future_map = {
                executor.submit(process_single_file, t): t["file_path"]
                for t in tasks
            }

            for future in as_completed(future_map):
                fp = future_map[future]
                try:
                    payload = future.result(timeout=WORKER_TIMEOUT_SEC)
                    all_results.extend(payload["results"])
                    all_queries.extend(payload["queries"])
                    all_errors.extend(payload["errors"])

                except FutureTimeoutError:
                    timed_out += 1
                    msg = (f"[WORKER TIMEOUT] {fp} did not finish within "
                           f"{WORKER_TIMEOUT_SEC}s — all its rows → TRANSFORM_NOT_FOUND")
                    all_errors.append(msg)
                    log.error(msg)
                    # Mark every pair in this task as not-found so the row
                    # still appears in the output with a clear sentinel.
                    task = next(t for t in tasks if t["file_path"] == fp)
                    for p in task["pairs"]:
                        all_results.append({
                            **p,
                            "transformation_logic": TRANSFORM_NF,
                            "query_uuid": "",
                        })

                except Exception as exc:
                    all_errors.append(f"[FUTURE ERROR] {fp}: {exc}")
                    log.error("Future failed for %s: %s", fp, exc)
                    # Worker crashed without returning results — mark every pair
                    # in this file as TRANSFORM_NOT_FOUND so the row still
                    # appears in the output and is not silently dropped.
                    task = next((t for t in tasks if t["file_path"] == fp), None)
                    if task:
                        for p in task["pairs"]:
                            all_results.append({
                                **p,
                                "transformation_logic": TRANSFORM_NF,
                                "query_uuid": "",
                            })

                completed += 1
                if completed % 100 == 0 or completed == len(tasks):
                    log.info("Progress: %d / %d files processed", completed, len(tasks))

    except KeyboardInterrupt:
        log.warning("KeyboardInterrupt received — shutting down workers and saving partial results …")
        # Fall through to write whatever was collected before interruption

    # ── Log non-fatal errors ──────────────────────────────────────────────────
    if all_errors:
        log.warning("%d non-fatal warnings encountered (see log):", len(all_errors))
        for e in all_errors[:20]:
            log.warning("  %s", e)
        if len(all_errors) > 20:
            log.warning("  … and %d more.", len(all_errors) - 20)

    # ── Reconstruct Fact Lineage DataFrame ────────────────────────────────────
    log.info("Reconstructing fact lineage dataframe …")
    results_index: dict[int, dict] = {r["row_id"]: r for r in all_results}

    new_rows = []
    for orig_row in df.itertuples(index=True):
        enriched = results_index.get(orig_row.Index, {})
        logic    = enriched.get("transformation_logic", TRANSFORM_NF)
        quuid    = enriched.get("query_uuid",           "")
        new_rows.append({
            "source_table":         getattr(orig_row, "source_table",  ""),
            "source_column":        getattr(orig_row, "source_column", ""),
            "target_table":         getattr(orig_row, "target_table",  ""),
            "target_column":        getattr(orig_row, "target_column", ""),
            "relation":             getattr(orig_row, "relation",      ""),
            # Use parent_original (the relative path from CSV), NOT the
            # resolved absolute path. Per requirement: "keep same value like
            # input csv, no need to mention full path from where it was read".
            "parent":               getattr(orig_row, "parent_original",
                                            getattr(orig_row, "parent", "")),
            "parent_type":          getattr(orig_row, "parent_type",   ""),
            "transformation_logic": logic,
            "query_uuid":           quuid,
        })

    fact_df = pd.DataFrame(new_rows)
    log.info("Fact dataframe ready: %d rows", len(fact_df))

    # ── Build Dim Queries DataFrame (de-duplicated on uuid) ───────────────────
    QUERY_COLS = ["query_uuid", "sql_query", "file_path", "parse_status", "parse_error"]
    if all_queries:
        queries_df = (
            pd.DataFrame(all_queries)
              .drop_duplicates(subset=["query_uuid"])
              .reindex(columns=QUERY_COLS)
              .reset_index(drop=True)
        )
    else:
        queries_df = pd.DataFrame(columns=QUERY_COLS)
        log.warning("all_queries is empty — every worker failed or produced no statements.")
    n_failed = int((queries_df["parse_status"] == "PARSE_FAILED").sum()) if not queries_df.empty else 0
    log.info(
        "Dim queries dataframe ready: %d unique statements (%d parse failures)",
        len(queries_df), n_failed,
    )
    if n_failed:
        log.warning(
            "%d statement(s) failed to parse — those rows carry "
            "transformation_logic=PARSE_FAILED in the Fact sheet; "
            "look up their query_uuid in Enterprise_Dim_Queries.xlsx "
            "to see the raw SQL and error.",
            n_failed,
        )
    if timed_out:
        log.warning("%d worker(s) timed out — increase --worker-timeout if needed.", timed_out)

    # ── Write Excel outputs ───────────────────────────────────────────────────
    fact_path    = str(out_dir / "Enterprise_Fact_Lineage.xlsx")
    queries_path = str(out_dir / "Enterprise_Dim_Queries.xlsx")

    write_fact_lineage_xlsx(fact_df,    fact_path)
    write_dim_queries_xlsx (queries_df, queries_path)

    log.info("=" * 70)
    log.info("Pipeline complete.")
    log.info("  Fact Lineage  → %s", fact_path)
    log.info("  Dim Queries   → %s", queries_path)
    if n_failed:
        log.info(
            "  ⚠  %d parse-failed statement(s) — filter Fact sheet on "
            "transformation_logic = PARSE_FAILED and cross-ref query_uuid "
            "in Dim Queries for the raw SQL + error.",
            n_failed,
        )
    log.info("=" * 70)


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Required on Windows so that spawned worker processes do not try to
    # re-import __main__ and re-launch the pool recursively.
    freeze_support()

    import argparse

    parser = argparse.ArgumentParser(
        description="Enterprise Column Lineage Transformation Extractor",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--input",
        required=True,
        metavar="INPUT_CSV",
        help="Path to the input lineage CSV file (~180 k rows).",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        metavar="OUTPUT_DIR",
        help="Directory where the two Excel files will be written.",
    )
    parser.add_argument(
        "--sql-base-dir",
        default=None,
        metavar="BASE_DIR",
        help=(
            "Root directory for resolving relative `parent` paths in the CSV. "
            "Defaults to the directory where this script lives, so paths like "
            "'ActiveInventory_Modified/.../file.sql' work with no extra flags. "
            "Override only if your SQL files live somewhere else entirely."
        ),
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=None,
        metavar="N",
        help="Number of parallel worker processes (default: cpu_count).",
    )
    parser.add_argument(
        "--worker-timeout",
        type=int,
        default=WORKER_TIMEOUT_SEC,
        metavar="SECONDS",
        help=(
            "Maximum seconds a single worker subprocess may run before it is "
            "cancelled and its rows marked TRANSFORM_NOT_FOUND. "
            f"Default: {WORKER_TIMEOUT_SEC}."
        ),
    )
    parser.add_argument(
        "--chunksize",
        type=int,
        default=4,
        metavar="N",
        help="Executor.map chunksize for batching tasks to workers.",
    )

    args = parser.parse_args()

    run_pipeline(
        input_csv      = args.input,
        output_dir     = args.output_dir,
        sql_base_dir   = args.sql_base_dir,
        max_workers    = args.workers,
        worker_timeout = args.worker_timeout,
        chunksize      = args.chunksize,
    )
