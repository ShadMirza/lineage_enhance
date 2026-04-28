from __future__ import annotations

import argparse
import logging
import os
import re
import signal
import sys
import threading
import uuid
from concurrent.futures import ProcessPoolExecutor, TimeoutError as FutureTimeoutError, as_completed
from dataclasses import dataclass
from multiprocessing import cpu_count, freeze_support
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlglot import exp, parse_one
from tqdm import tqdm

sys.setrecursionlimit(5000)

LOGGER = logging.getLogger("enterprise_lineage_extractor")

MAX_STMTS_PER_FILE = 500
MAX_STMT_LENGTH_CHARS = 200000
DEFAULT_STMT_TIMEOUT = 30

SAFE_BLOCK_COMMENT_RE = re.compile(r"/\*[^*]*\*+(?:[^*/][^*]*\*+)*/")
LINE_COMMENT_RE = re.compile(r"--[^\n]*")
CONTROL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

STATIC_MARKERS = {
    "*",
    "STATIC",
    "STATICVALUE",
    "STATIC VALUE",
    "STATIC_VALUE",
    "HARDCODED",
    "HARDCODED VALUE",
    "HARDCODED_VALUE",
    "LITERAL",
    "LITERAL VALUE",
    "LITERAL_VALUE",
    "CONSTANT",
    "CONSTANT VALUE",
    "CONSTANT_VALUE",
    "DERIVED",
}

TRANSFORM_NOT_FOUND = "TRANSFORM_NOT_FOUND"
STATIC_FALLBACK = "STATIC_FALLBACK"
DIRECT_PASSTHRU = "Direct pass through"


@dataclass
class StatementInfo:
    uuid: str
    parsed: Optional[exp.Expression]
    restored: str
    parse_ok: bool
    parse_error: str
    target_table: str


def _sanitize_value(v: object) -> str:
    if v is None:
        return ""
    text = str(v)
    if text.lower() == "nan":
        return ""
    return CONTROL_CHARS_RE.sub("", text)


def _norm(v: str) -> str:
    return _sanitize_value(v).strip().upper()


def _strip_qualifiers(col_ref: str) -> str:
    text = _norm(col_ref).replace('"', "")
    parts = [p for p in text.split(".") if p]
    return parts[-1] if parts else ""


def _bare_table(ref: str) -> str:
    text = _norm(ref).replace('"', "")
    text = re.sub(r"^\$\{?[A-Z0-9_]+\}?\.", "", text)
    parts = [p for p in text.split(".") if p]
    if len(parts) >= 1:
        return parts[-1]
    return ""


def _table_matches(csv_table: str, sql_table: str) -> bool:
    a = _bare_table(csv_table)
    b = _bare_table(sql_table)
    return bool(a and b and a == b)


def _replace_shell_vars(sql_text: str) -> Tuple[str, Dict[str, str]]:
    mapping: Dict[str, str] = {}

    def _repl(match: re.Match) -> str:
        key = f"__SHELLVAR_{len(mapping)}__"
        mapping[key] = match.group(0)
        return key

    pattern = re.compile(r"\$\{?[A-Za-z_][A-Za-z0-9_]*\}?")
    replaced = pattern.sub(_repl, sql_text)
    return replaced, mapping


def _restore_shell_vars(sql_text: str, mapping: Dict[str, str]) -> str:
    out = sql_text
    for k, v in mapping.items():
        out = out.replace(k, v)
    return out


def _remove_comments(sql_text: str) -> str:
    no_block = SAFE_BLOCK_COMMENT_RE.sub(" ", sql_text)
    no_line = LINE_COMMENT_RE.sub(" ", no_block)
    return no_line


def _split_statements(sql_text: str) -> List[str]:
    cleaned = _remove_comments(sql_text)
    parts = [p.strip() for p in cleaned.split(";")]
    return [p for p in parts if p]


def _is_pure_ddl(stmt: str) -> bool:
    s = _norm(stmt)
    if re.match(r"^DROP\s+TABLE\b", s):
        return True
    if re.match(r"^ALTER\s+TABLE\b", s) or re.match(r"^ALTER\s+COLUMN\b", s):
        return True
    if re.match(r"^COLLECT\s+STATISTICS\b", s):
        return True
    if re.match(r"^CREATE\s+(MULTISET\s+|VOLATILE\s+|GLOBAL\s+TEMPORARY\s+)?TABLE\b", s) and " AS SELECT " not in f" {s} ":
        return True
    return False


def _extract_select_body_for_ctas(stmt: str) -> str:
    pattern = re.compile(
        r"^\s*(?:CREATE\s+(?:MULTISET\s+|VOLATILE\s+|GLOBAL\s+TEMPORARY\s+)?TABLE\s+.+?|REPLACE\s+(?:RECURSIVE\s+)?VIEW\s+.+?)(?=\s+SELECT\b)",
        re.IGNORECASE | re.DOTALL,
    )
    if re.search(r"\bAS\s+SELECT\b", stmt, flags=re.IGNORECASE):
        idx = re.search(r"\bSELECT\b", stmt, flags=re.IGNORECASE)
        if idx:
            return stmt[idx.start():]
    if re.match(r"\s*REPLACE\s+(?:RECURSIVE\s+)?VIEW\b", stmt, flags=re.IGNORECASE):
        idx = re.search(r"\bSELECT\b", stmt, flags=re.IGNORECASE)
        if idx:
            return stmt[idx.start():]
    return pattern.sub("", stmt).strip()


def _parse_with_timeout(stmt: str, timeout_seconds: int) -> exp.Expression:
    if os.name != "nt":
        def _handler(signum, frame):
            raise TimeoutError("Statement parse timeout")

        previous = signal.signal(signal.SIGALRM, _handler)
        signal.alarm(timeout_seconds)
        try:
            return parse_one(stmt, read="teradata")
        finally:
            signal.alarm(0)
            signal.signal(signal.SIGALRM, previous)

    holder: Dict[str, object] = {}

    def _runner() -> None:
        try:
            holder["result"] = parse_one(stmt, read="teradata")
        except Exception as exc:  # noqa: BLE001
            holder["error"] = exc

    t = threading.Thread(target=_runner, daemon=True)
    t.start()
    t.join(timeout_seconds)
    if t.is_alive():
        raise TimeoutError("Statement parse timeout")
    if "error" in holder:
        raise holder["error"]
    return holder["result"]


def _statement_target_table(parsed: exp.Expression) -> str:
    if isinstance(parsed, exp.Insert):
        target = parsed.this
        if isinstance(target, exp.Schema):
            target = target.this
        if isinstance(target, exp.Table):
            return _bare_table(target.name)
    elif isinstance(parsed, exp.Update):
        if isinstance(parsed.this, exp.Table):
            return _bare_table(parsed.this.name)
    elif isinstance(parsed, exp.Merge):
        if isinstance(parsed.this, exp.Table):
            return _bare_table(parsed.this.name)
    return ""


def _resolve_source_alias(parsed: exp.Expression, src_tbl: str) -> str:
    src_tbl_n = _bare_table(src_tbl)
    for tbl_node in parsed.find_all(exp.Table):
        if _bare_table(tbl_node.name) == src_tbl_n:
            alias = str(tbl_node.alias).strip() if tbl_node.alias else ""
            return _norm(alias) or src_tbl_n
    return ""


def _extract_from_select(parsed: exp.Expression, src_col: str, tgt_col: str) -> List[str]:
    out: List[str] = []
    tgt = _strip_qualifiers(tgt_col)
    src = _strip_qualifiers(src_col)

    for sel in parsed.find_all(exp.Select):
        for projection in sel.expressions or []:
            alias = _norm(projection.alias_or_name or "")
            implicit = _norm(_strip_qualifiers(projection.sql(dialect="teradata")))
            if tgt and (alias == tgt or implicit == tgt):
                out.append(projection.sql(dialect="teradata"))
            elif src and src in _norm(projection.sql(dialect="teradata")):
                out.append(projection.sql(dialect="teradata"))

    seen = set()
    dedup = []
    for x in out:
        if x not in seen:
            seen.add(x)
            dedup.append(x)
    return dedup


def _extract_from_insert(parsed: exp.Expression, src_col: str, tgt_col: str) -> List[str]:
    out: List[str] = []
    if not isinstance(parsed, exp.Insert):
        return out

    tgt = _strip_qualifiers(tgt_col)
    src = _strip_qualifiers(src_col)
    cols = []
    if isinstance(parsed.this, exp.Schema):
        cols = [_strip_qualifiers(c.sql()) for c in parsed.this.expressions or []]

    expr = parsed.expression
    if isinstance(expr, exp.Values):
        for tup in expr.find_all(exp.Tuple):
            values = tup.expressions or []
            if tgt and cols and tgt in [_norm(c) for c in cols]:
                idx = [_norm(c) for c in cols].index(tgt)
                if idx < len(values):
                    out.append(values[idx].sql(dialect="teradata"))
            elif not tgt:
                out.extend(v.sql(dialect="teradata") for v in values)
        return list(dict.fromkeys(out))

    sel_vals = _extract_from_select(parsed, src, tgt)
    out.extend(sel_vals)
    return list(dict.fromkeys(out))


def _extract_from_update(parsed: exp.Expression, src_col: str, tgt_col: str) -> List[str]:
    out: List[str] = []
    if not isinstance(parsed, exp.Update):
        return out
    tgt = _strip_qualifiers(tgt_col)
    for eq in parsed.find_all(exp.EQ):
        left = _strip_qualifiers(eq.left.sql())
        if tgt and _norm(left) == _norm(tgt):
            out.append(eq.right.sql(dialect="teradata"))
    return list(dict.fromkeys(out))


def _extract_from_merge(parsed: exp.Expression, src_col: str, tgt_col: str) -> List[str]:
    out: List[str] = []
    if not isinstance(parsed, exp.Merge):
        return out

    tgt = _strip_qualifiers(tgt_col)

    for when in parsed.find_all(exp.When):
        for eq in when.find_all(exp.EQ):
            left = _strip_qualifiers(eq.left.sql())
            if tgt and _norm(left) == _norm(tgt):
                out.append(eq.right.sql(dialect="teradata"))

        ins = when.find(exp.Insert)
        if ins and isinstance(ins.expression, exp.Values):
            cols = [_strip_qualifiers(c.sql()) for c in (ins.this.expressions or [])] if isinstance(ins.this, exp.Schema) else []
            for tup in ins.expression.find_all(exp.Tuple):
                vals = tup.expressions or []
                if tgt and cols and tgt in [_norm(c) for c in cols]:
                    idx = [_norm(c) for c in cols].index(tgt)
                    if idx < len(vals):
                        out.append(vals[idx].sql(dialect="teradata"))

    return list(dict.fromkeys(out))


def _extract_static_target_logic(parsed: exp.Expression, tgt_col: str) -> List[str]:
    out: List[str] = []
    out.extend(_extract_from_update(parsed, "", tgt_col))
    out.extend(_extract_from_insert(parsed, "", tgt_col))
    out.extend(_extract_from_merge(parsed, "", tgt_col))
    out.extend(_extract_from_select(parsed, "", tgt_col))
    return list(dict.fromkeys([x for x in out if x]))


def _is_literal_expression(expr_sql: str) -> bool:
    s = _norm(expr_sql)
    if s == "NULL":
        return True
    if s.startswith("'") and s.endswith("'"):
        return True
    if re.match(r"^[+-]?\d+(\.\d+)?$", s):
        return True
    if s in {"CURRENT_DATE", "CURRENT_TIMESTAMP", "CURRENT_TIME"}:
        return True
    return False


def _is_passthrough(expr_sql: str, src_col: str, src_alias: Optional[str] = None) -> bool:
    bare = _strip_qualifiers(expr_sql)
    src = _strip_qualifiers(src_col)
    if not bare or not src or _norm(bare) != _norm(src):
        return False

    if src_alias:
        expr_norm = _norm(expr_sql)
        if "." in expr_norm:
            qualifier = expr_norm.split(".")[-2]
            if qualifier not in {_norm(src_alias), _norm(_bare_table(src_alias))}:
                return False
    return True


def extract_transformation(parsed: exp.Expression, src_col: str, tgt_col: str, is_static: bool, src_alias: str) -> List[str]:
    raw_logics: List[str] = []
    if isinstance(parsed, exp.Merge):
        raw_logics = _extract_from_merge(parsed, src_col, tgt_col)
    elif isinstance(parsed, exp.Update):
        raw_logics = _extract_from_update(parsed, src_col, tgt_col)
    elif isinstance(parsed, exp.Insert):
        raw_logics = _extract_from_insert(parsed, src_col, tgt_col)
    elif isinstance(parsed, exp.Select):
        raw_logics = _extract_from_select(parsed, src_col, tgt_col)

    if isinstance(raw_logics, str):
        raw_logics = [raw_logics] if raw_logics not in (TRANSFORM_NOT_FOUND, STATIC_FALLBACK, "") else []
    elif not isinstance(raw_logics, list):
        raw_logics = []

    out: List[str] = []
    for logic in raw_logics:
        if _is_passthrough(logic, src_col, src_alias):
            out.append(DIRECT_PASSTHRU)
        else:
            out.append(logic)
    return list(dict.fromkeys(out))


def _extract_read_context(parsed: exp.Expression, src_col: str) -> List[str]:
    out: List[str] = []
    src = _strip_qualifiers(src_col)
    if not src:
        return out

    for sel in parsed.find_all(exp.Select):
        for proj in sel.expressions or []:
            proj_sql = proj.sql(dialect="teradata")
            if _norm(src) in _norm(proj_sql):
                out.append(f"Read in SELECT projection: {proj_sql}")

    for join in parsed.find_all(exp.Join):
        if join.args.get("on") is not None:
            pred = join.args["on"].sql(dialect="teradata")
            if _norm(src) in _norm(pred):
                out.append(f"Read in JOIN ON condition: {pred}")

    where = parsed.args.get("where")
    if where is not None:
        pred = where.sql(dialect="teradata")
        if _norm(src) in _norm(pred):
            out.append(f"Read in WHERE filter: {pred}")

    group = parsed.args.get("group")
    if group is not None and _norm(src) in _norm(group.sql(dialect="teradata")):
        out.append("Read in GROUP BY clause")

    order = parsed.args.get("order")
    if order is not None and _norm(src) in _norm(order.sql(dialect="teradata")):
        out.append("Read in ORDER BY clause")

    return list(dict.fromkeys(out))


def _candidate_score(
    stmt_target_table: str,
    tgt_table: str,
    expr_sql: str,
    src_table: str,
    src_alias: str,
    parsed: exp.Expression,
) -> int:
    targets_match = _table_matches(tgt_table, stmt_target_table)
    expr_norm = _norm(expr_sql)

    alias_ok = True
    if expr_sql != DIRECT_PASSTHRU and "." in expr_norm:
        parts = [p for p in expr_norm.split(".") if p]
        qual = parts[-2] if len(parts) >= 2 else ""
        if src_alias:
            if qual not in {_norm(src_alias), _norm(_bare_table(src_table))}:
                alias_ok = False
        else:
            table_aliases = set()
            for t in parsed.find_all(exp.Table):
                if t.alias:
                    table_aliases.add(_norm(str(t.alias)))
            if qual in table_aliases:
                src_tbl_norm = _norm(_bare_table(src_table))
                if qual != src_tbl_norm:
                    alias_ok = False

    if not alias_ok:
        return 0

    src_ref = _norm(_bare_table(src_table))
    src_hit = (src_ref and src_ref in expr_norm) or (src_alias and _norm(src_alias) in expr_norm)

    if targets_match and src_hit:
        return 3
    if targets_match:
        return 2
    return 1


def _build_statement_catalog_for_file(
    file_path: Path,
    file_path_original: str,
    stmt_timeout: int,
) -> List[StatementInfo]:
    out: List[StatementInfo] = []

    if not file_path.exists():
        LOGGER.warning("PARSE_WARN missing file: %s", file_path)
        uid = str(uuid.uuid4())
        out.append(StatementInfo(uid, None, "", False, "FILE_NOT_FOUND", ""))
        return out

    sql_text = file_path.read_text(encoding="utf-8", errors="ignore")
    replaced, mapping = _replace_shell_vars(sql_text)
    statements = _split_statements(replaced)[:MAX_STMTS_PER_FILE]

    for stmt in statements:
        restored = _restore_shell_vars(stmt, mapping)
        qid = str(uuid.uuid4())

        if len(stmt) > MAX_STMT_LENGTH_CHARS:
            LOGGER.warning("SIZE_SKIP statement too large in %s", file_path_original)
            out.append(StatementInfo(qid, None, restored, False, "STATEMENT_TOO_LARGE", ""))
            continue

        if _is_pure_ddl(stmt):
            out.append(StatementInfo(qid, None, restored, False, "PURE_DDL_SKIPPED", ""))
            continue

        parse_input = stmt
        s_norm = _norm(stmt)
        if s_norm.startswith("CREATE ") or s_norm.startswith("REPLACE "):
            parse_input = _extract_select_body_for_ctas(stmt)

        try:
            parsed = _parse_with_timeout(parse_input, stmt_timeout)
            if isinstance(parsed, (exp.Drop, exp.Alter, exp.Command)):
                out.append(StatementInfo(qid, None, restored, False, "DDL_NODE_SKIPPED", ""))
                continue
            if isinstance(parsed, exp.Create) and not parsed.expression:
                out.append(StatementInfo(qid, None, restored, False, "CREATE_NO_SELECT_SKIPPED", ""))
                continue

            out.append(StatementInfo(qid, parsed, restored, True, "", _statement_target_table(parsed)))
        except TimeoutError:
            LOGGER.warning("TIMEOUT parse timeout in %s", file_path_original)
            out.append(StatementInfo(qid, None, restored, False, "PARSE_TIMEOUT", ""))
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("PARSE_WARN parse failed in %s: %s", file_path_original, exc)
            out.append(StatementInfo(qid, None, restored, False, str(exc)[:2000], ""))

    if not out:
        qid = str(uuid.uuid4())
        out.append(StatementInfo(qid, None, "", False, "NO_STATEMENTS", ""))

    return out


def _default_uuid(statements: Sequence[StatementInfo]) -> str:
    for s in statements:
        if s.parse_ok:
            return s.uuid
    return statements[0].uuid if statements else ""


def _row_to_outputs(row: Dict[str, str], statements: Sequence[StatementInfo]) -> List[Dict[str, str]]:
    src_table = _sanitize_value(row.get("source_table", ""))
    src_col = _sanitize_value(row.get("source_column", ""))
    tgt_table = _sanitize_value(row.get("target_table", ""))
    tgt_col = _sanitize_value(row.get("target_column", ""))
    relation = _sanitize_value(row.get("relation", ""))
    parent = _sanitize_value(row.get("parent", ""))
    parent_type = _sanitize_value(row.get("parent_type", ""))

    src_col_n = _norm(src_col)
    tgt_col_n = _norm(tgt_col)
    default_uuid = _default_uuid(statements)

    base = {
        "source_table": src_table,
        "source_column": src_col,
        "target_table": tgt_table,
        "target_column": tgt_col,
        "relation": relation,
        "parent": parent,
        "parent_type": parent_type,
    }

    # GATE 1
    if src_col_n == "NULL" or tgt_col_n == "NULL":
        for s in statements:
            if s.parse_ok:
                vals = _extract_static_target_logic(s.parsed, tgt_col)
                if vals:
                    return [{**base, "transformation_logic": "NULL", "query_uuid": s.uuid}]
        return [{**base, "transformation_logic": "NULL", "query_uuid": default_uuid}]

    # GATE 2
    if src_col_n == "*" or tgt_col_n == "*":
        text = "Wildcard SELECT * - all source columns"
        pick = default_uuid
        for s in statements:
            if s.parse_ok and _table_matches(tgt_table, s.target_table):
                pick = s.uuid
                break
        return [{**base, "transformation_logic": text, "query_uuid": pick}]

    # GATE 3
    if _norm(relation) == "READ" and src_col_n and not tgt_col_n:
        rows: List[Dict[str, str]] = []
        for s in statements:
            if not s.parse_ok:
                continue
            ctx = _extract_read_context(s.parsed, src_col)
            for c in ctx:
                rows.append({**base, "transformation_logic": c, "query_uuid": s.uuid})
        if rows:
            return rows

    # GATE 4
    if not src_col_n and not tgt_col_n:
        return [{**base, "transformation_logic": "MISSING_BOTH_COLUMNS", "query_uuid": default_uuid}]
    if not src_col_n:
        return [{**base, "transformation_logic": "MISSING_SOURCE_COLUMN", "query_uuid": default_uuid}]
    if not tgt_col_n:
        return [{**base, "transformation_logic": "MISSING_TARGET_COLUMN", "query_uuid": default_uuid}]

    # GATE 5
    is_static = src_col_n in STATIC_MARKERS
    if is_static:
        all_items: List[Tuple[str, str]] = []
        for s in statements:
            if not s.parse_ok:
                continue
            vals = _extract_static_target_logic(s.parsed, tgt_col)
            vals = [v for v in vals if v]
            for v in vals:
                all_items.append((v, s.uuid))

        if all_items:
            scored = []
            for val, uid in all_items:
                scored.append((2 if _is_literal_expression(val) else 1, val, uid))
            top_score = max(x[0] for x in scored)
            top = [x for x in scored if x[0] == top_score]
            lit_top = [x for x in top if _is_literal_expression(x[1])]
            top = lit_top or top
            out = []
            seen = set()
            for _, v, uid in top:
                k = (v, uid)
                if k in seen:
                    continue
                seen.add(k)
                out.append({**base, "transformation_logic": v, "query_uuid": uid})
            return out

    # GATE 6
    candidates: List[Tuple[int, str, str]] = []
    failed = [s for s in statements if not s.parse_ok]

    for s in statements:
        if not s.parse_ok:
            continue
        src_alias = _resolve_source_alias(s.parsed, src_table)
        logics = extract_transformation(s.parsed, src_col, tgt_col, is_static, src_alias)
        for logic in logics:
            score = _candidate_score(s.target_table, tgt_table, logic, src_table, src_alias, s.parsed)
            if score > 0:
                candidates.append((score, logic, s.uuid))

    if candidates:
        best = max(c[0] for c in candidates)
        tier = [(l, u) for sc, l, u in candidates if sc == best]
        seen = set()
        out = []
        for logic, uid in tier:
            key = (logic, uid)
            if key in seen:
                continue
            seen.add(key)
            out.append({**base, "transformation_logic": logic, "query_uuid": uid})
        return out

    # POST-SCAN FALLBACK
    if is_static:
        for s in statements:
            if s.parse_ok:
                vals = _extract_static_target_logic(s.parsed, tgt_col)
                if vals:
                    return [{**base, "transformation_logic": vals[0], "query_uuid": s.uuid}]
        return [{**base, "transformation_logic": STATIC_FALLBACK, "query_uuid": default_uuid}]

    if failed:
        return [{**base, "transformation_logic": "PARSE_FAILED", "query_uuid": failed[0].uuid}]

    return [{**base, "transformation_logic": TRANSFORM_NOT_FOUND, "query_uuid": default_uuid}]


def _process_parent_group(
    payload: Dict[str, object],
    sql_base_dir: str,
    stmt_timeout: int,
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], List[str]]:
    parent = str(payload["parent"])
    rows = payload["rows"]
    resolved = (Path(sql_base_dir) / parent).resolve()
    warnings: List[str] = []
    if not resolved.exists():
        warnings.append(f"Missing SQL file: {resolved}")

    statements = _build_statement_catalog_for_file(resolved, parent, stmt_timeout)

    dim_rows: List[Dict[str, str]] = []
    for s in statements:
        dim_rows.append(
            {
                "query_uuid": s.uuid,
                "sql_query": _sanitize_value(s.restored),
                "file_path": parent,
                "parse_status": "OK" if s.parse_ok else "PARSE_FAILED",
                "parse_error": _sanitize_value(s.parse_error),
            }
        )

    fact_rows: List[Dict[str, str]] = []
    for r in rows:
        fact_rows.extend(_row_to_outputs(r, statements))

    return fact_rows, dim_rows, warnings


def _setup_logging(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    LOGGER.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s")

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    fh = logging.FileHandler(output_dir / "lineage_extractor.log", encoding="utf-8")
    fh.setFormatter(fmt)

    LOGGER.handlers.clear()
    LOGGER.addHandler(sh)
    LOGGER.addHandler(fh)


def _apply_sheet_style(ws, kind: str) -> None:
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    fills = {
        "DIRECT_PASSTHRU": PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid"),
        "PARSE_FAILED": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        "MISSING": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "NULL_LITERAL": PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid"),
        "STATIC": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "FOUND": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "NF": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        "ALT": PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid"),
    }

    headers = [c.value for c in ws[1]]
    logic_idx = headers.index("transformation_logic") + 1 if "transformation_logic" in headers else None
    status_idx = headers.index("parse_status") + 1 if "parse_status" in headers else None

    for row_idx in range(2, ws.max_row + 1):
        row_fill = fills["ALT"] if row_idx % 2 == 0 else None
        semantic = None

        if kind == "fact" and logic_idx:
            logic = _norm(ws.cell(row=row_idx, column=logic_idx).value or "")
            if logic == _norm(DIRECT_PASSTHRU):
                semantic = fills["DIRECT_PASSTHRU"]
            elif logic == "PARSE_FAILED":
                semantic = fills["PARSE_FAILED"]
            elif logic.startswith("MISSING"):
                semantic = fills["MISSING"]
            elif logic == "NULL":
                semantic = fills["NULL_LITERAL"]
            elif logic in {_norm(STATIC_FALLBACK)} or "STATIC" in logic:
                semantic = fills["STATIC"]
            elif logic == _norm(TRANSFORM_NOT_FOUND):
                semantic = fills["NF"]
            else:
                semantic = fills["FOUND"]

        if kind == "dim" and status_idx:
            status = _norm(ws.cell(row=row_idx, column=status_idx).value or "")
            if status == "PARSE_FAILED":
                semantic = fills["PARSE_FAILED"]

        final_fill = semantic or row_fill
        if final_fill:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = final_fill

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            l = len(str(val)) if val is not None else 0
            max_len = max(max_len, l)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)


def _write_fact_workbook(path: Path, fact_df: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fact Lineage"

    cols = [
        "source_table",
        "source_column",
        "target_table",
        "target_column",
        "relation",
        "parent",
        "parent_type",
        "transformation_logic",
        "query_uuid",
    ]

    ws.append(cols)
    for _, row in fact_df[cols].iterrows():
        ws.append([_sanitize_value(row.get(c, "")) for c in cols])

    _apply_sheet_style(ws, "fact")

    sum_ws = wb.create_sheet("Summary")
    sum_ws.append(["Category", "Count"])
    summary = fact_df["transformation_logic"].fillna("").map(lambda x: _sanitize_value(x)).value_counts(dropna=False)
    for k, v in summary.items():
        sum_ws.append([k, int(v)])
    _apply_sheet_style(sum_ws, "summary")

    wb.save(path)


def _write_dim_workbook(path: Path, dim_df: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dim Queries"
    cols = ["query_uuid", "sql_query", "file_path", "parse_status", "parse_error"]
    ws.append(cols)
    for _, row in dim_df[cols].iterrows():
        ws.append([_sanitize_value(row.get(c, "")) for c in cols])
    _apply_sheet_style(ws, "dim")
    wb.save(path)


def run_pipeline(
    input_csv: Path,
    output_dir: Path,
    sql_base_dir: Path,
    workers: int,
    worker_timeout: int,
    chunksize: int,
    stmt_timeout: int = DEFAULT_STMT_TIMEOUT,
) -> None:
    _setup_logging(output_dir)

    df = pd.read_csv(input_csv)
    required = ["source_table", "source_column", "target_table", "target_column", "relation", "parent", "parent_type"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    for c in required:
        df[c] = df[c].map(_sanitize_value)

    df = df.drop_duplicates(subset=["source_table", "source_column", "target_table", "target_column", "parent"]).reset_index(drop=True)

    grouped = []
    for parent, g in df.groupby("parent", dropna=False):
        grouped.append(
            {
                "parent": _sanitize_value(parent),
                "rows": g[required].to_dict("records"),
            }
        )

    LOGGER.info("Total grouped SQL files: %d", len(grouped))

    all_fact: List[Dict[str, str]] = []
    all_dim: List[Dict[str, str]] = []

    with ProcessPoolExecutor(max_workers=workers) as ex:
        future_map = {
            ex.submit(_process_parent_group, payload, str(sql_base_dir), stmt_timeout): payload["parent"]
            for payload in grouped
        }

        try:
            for idx, fut in enumerate(tqdm(as_completed(future_map), total=len(future_map), desc="Processing SQL files"), start=1):
                parent = future_map[fut]
                try:
                    fact_rows, dim_rows, warns = fut.result(timeout=worker_timeout)
                    all_fact.extend(fact_rows)
                    all_dim.extend(dim_rows)
                    for w in warns:
                        LOGGER.warning("EXTRACT_WARN %s", w)
                except FutureTimeoutError:
                    LOGGER.error("FUTURE_ERROR worker timeout for parent: %s", parent)
                    subset = df[df["parent"] == parent]
                    for _, r in subset.iterrows():
                        all_fact.append(
                            {
                                "source_table": r["source_table"],
                                "source_column": r["source_column"],
                                "target_table": r["target_table"],
                                "target_column": r["target_column"],
                                "relation": r["relation"],
                                "parent": r["parent"],
                                "parent_type": r["parent_type"],
                                "transformation_logic": TRANSFORM_NOT_FOUND,
                                "query_uuid": "",
                            }
                        )
                except Exception as exc:  # noqa: BLE001
                    LOGGER.error("FUTURE_ERROR parent=%s error=%s", parent, exc)

                if idx % 100 == 0 or idx == len(future_map):
                    LOGGER.info("Progress: %d / %d files processed", idx, len(future_map))

        except KeyboardInterrupt:
            LOGGER.warning("KeyboardInterrupt detected. Shutting down gracefully and writing partial outputs.")
            ex.shutdown(wait=False, cancel_futures=True)

    fact_df = pd.DataFrame(
        all_fact,
        columns=[
            "source_table",
            "source_column",
            "target_table",
            "target_column",
            "relation",
            "parent",
            "parent_type",
            "transformation_logic",
            "query_uuid",
        ],
    )

    dim_df = pd.DataFrame(
        all_dim,
        columns=["query_uuid", "sql_query", "file_path", "parse_status", "parse_error"],
    )

    if dim_df.empty:
        dim_df = pd.DataFrame(columns=["query_uuid", "sql_query", "file_path", "parse_status", "parse_error"])

    fact_out = output_dir / "Enterprise_Fact_Lineage.xlsx"
    dim_out = output_dir / "Enterprise_Dim_Queries.xlsx"

    _write_fact_workbook(fact_out, fact_df)
    _write_dim_workbook(dim_out, dim_df)

    LOGGER.info("Fact output written: %s", fact_out)
    LOGGER.info("Dim output written: %s", dim_out)
    LOGGER.warning("For PARSE_FAILED entries, cross-reference query_uuid in Enterprise_Dim_Queries.xlsx")


def parse_args(argv: Optional[Iterable[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Enterprise Column Lineage Transformation Extractor")
    parser.add_argument("--input", required=True, help="Path to input CSV")
    parser.add_argument("--output-dir", default=".", help="Directory for output Excel files")
    parser.add_argument("--sql-base-dir", default=str(Path(__file__).resolve().parent), help="Base dir for parent SQL paths")
    parser.add_argument("--workers", type=int, default=cpu_count(), help="Process pool workers")
    parser.add_argument("--worker-timeout", type=int, default=300, help="Worker timeout seconds")
    parser.add_argument("--chunksize", type=int, default=4, help="Executor chunksize hint")
    parser.add_argument("--statement-timeout", type=int, default=DEFAULT_STMT_TIMEOUT, help="Per statement parse timeout")
    return parser.parse_args(argv)


def main(argv: Optional[Iterable[str]] = None) -> int:
    args = parse_args(argv)
    run_pipeline(
        input_csv=Path(args.input),
        output_dir=Path(args.output_dir),
        sql_base_dir=Path(args.sql_base_dir),
        workers=max(1, int(args.workers)),
        worker_timeout=max(1, int(args.worker_timeout)),
        chunksize=max(1, int(args.chunksize)),
        stmt_timeout=max(1, int(args.statement_timeout)),
    )
    return 0


if __name__ == "__main__":
    freeze_support()
    raise SystemExit(main())
